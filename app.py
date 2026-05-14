# =====================================================
# IMPORTS GENERALES
# =====================================================
import os
import sqlite3
import bcrypt
import zipfile
import tempfile
import shutil
import click
import secrets
from datetime import datetime, timedelta
from io import BytesIO

from dotenv import load_dotenv

from flask import (
    Flask, render_template, request,
    redirect, url_for, session, flash, g, send_file
)

from flask_wtf.csrf import CSRFProtect, CSRFError
from werkzeug.utils import secure_filename
from werkzeug.exceptions import RequestEntityTooLarge

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# =====================================================
# CARGA DE VARIABLES DE ENTORNO
# =====================================================
load_dotenv()


# =====================================================
# CONFIGURACIÓN PRINCIPAL DE FLASK
# =====================================================
app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY")

if not app.secret_key:
    raise RuntimeError("Falta configurar SECRET_KEY en el archivo .env o variables de entorno.")

# =====================================================
# FILTRO JINJA: FORMATO MONEDA COLONES SIN DECIMALES
# Uso en HTML: {{ monto|crc }}
# Ejemplo: 825.0 -> ₡825
# =====================================================
@app.template_filter("crc")
def format_crc(value):
    try:
        return f"₡{float(value):,.0f}"
    except (TypeError, ValueError):
        return "₡0"

# =====================================================
# PROTECCIÓN CSRF
# =====================================================
csrf = CSRFProtect(app)


@app.errorhandler(CSRFError)
def handle_csrf_error(e):
    flash("La sesión del formulario expiró o no es válida. Intenta nuevamente.", "warning")
    return redirect(request.referrer or url_for("inicio"))


# =====================================================
# BASE DE DATOS
# =====================================================
database_path = os.environ.get("DATABASE_PATH", os.path.join(app.instance_path, "aserve.db"))

if not os.path.isabs(database_path):
    database_path = os.path.join(os.getcwd(), database_path)

app.config["DATABASE"] = database_path


# =====================================================
# SUBIDA DE IMÁGENES
# =====================================================
upload_folder = os.environ.get("UPLOAD_FOLDER", os.path.join("static", "uploads"))
app.config["UPLOAD_FOLDER"] = upload_folder
app.config["ALLOWED_EXTENSIONS"] = {"png", "jpg", "jpeg", "webp"}

# Límite máximo de subida: 2 MB
app.config["MAX_CONTENT_LENGTH"] = 2 * 1024 * 1024

# =====================================================
# ERROR: IMAGEN DEMASIADO PESADA
# =====================================================
@app.errorhandler(RequestEntityTooLarge)
def handle_file_too_large(e):
    flash("El archivo es demasiado pesado. La imagen no debe superar los 2 MB.", "warning")
    return redirect(request.referrer or url_for("admin_products"))

# =====================================================
# SEGURIDAD DE COOKIES
# =====================================================
app.config["SESSION_COOKIE_HTTPONLY"] = True
app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = os.environ.get("SESSION_COOKIE_SECURE", "0") == "1"


# =====================================================
# FUNCIONES DE BASE DE DATOS
# =====================================================
def get_db():
    if "db" not in g:
        db_folder = os.path.dirname(app.config["DATABASE"])
        os.makedirs(db_folder, exist_ok=True)

        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row

    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    db = g.pop("db", None)

    if db is not None:
        db.close()


# =====================================================
# FUNCIÓN AUXILIAR: ASEGURAR ESTRUCTURA DE BASE DE DATOS
# - Sirve para bases ya existentes sin borrar información
# - Crea audit_logs si no existe
# - Agrega columnas nuevas si faltan
# - Crea índices recomendados si no existen
# =====================================================
def ensure_app_schema(db):
    try:
        # =====================================================
        # TABLA: audit_logs
        # Guarda acciones administrativas importantes
        # =====================================================
        db.execute("""
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                fecha TEXT NOT NULL,
                admin_id INTEGER,
                admin_nombre TEXT NOT NULL DEFAULT 'Sistema',
                accion TEXT NOT NULL,
                detalle TEXT
            )
        """)

        # =====================================================
        # REVISAR COLUMNAS DE USERS
        # =====================================================
        columnas_users = db.execute("PRAGMA table_info(users)").fetchall()
        columnas_users = [c["name"] for c in columnas_users]

        if "password_temporal" not in columnas_users:
            db.execute("""
                ALTER TABLE users
                ADD COLUMN password_temporal INTEGER NOT NULL DEFAULT 0
            """)

        # =====================================================
        # REVISAR COLUMNAS DE PRODUCTS
        # Esto ayuda si tu base vieja no tenía imagen o stock mínimo.
        # =====================================================
        columnas_products = db.execute("PRAGMA table_info(products)").fetchall()
        columnas_products = [c["name"] for c in columnas_products]

        if "stock_minimo" not in columnas_products:
            db.execute("""
                ALTER TABLE products
                ADD COLUMN stock_minimo INTEGER NOT NULL DEFAULT 0
            """)

        if "image_filename" not in columnas_products:
            db.execute("""
                ALTER TABLE products
                ADD COLUMN image_filename TEXT
            """)

        if "activo" not in columnas_products:
            db.execute("""
                ALTER TABLE products
                ADD COLUMN activo INTEGER NOT NULL DEFAULT 1
            """)

        # =====================================================
        # ÍNDICES RECOMENDADOS
        # IF NOT EXISTS evita errores si ya fueron creados.
        # =====================================================

        # Usuarios
        db.execute("CREATE INDEX IF NOT EXISTS idx_users_usuario ON users(usuario)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_users_estado ON users(estado)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_users_rol ON users(rol)")

        # Productos
        db.execute("CREATE INDEX IF NOT EXISTS idx_products_nombre ON products(nombre)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_products_activo ON products(activo)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_products_stock ON products(stock)")

        # Órdenes
        db.execute("CREATE INDEX IF NOT EXISTS idx_orders_fecha ON orders(fecha)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_orders_user_id ON orders(user_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_orders_tipo_pago ON orders(tipo_pago)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_orders_estado ON orders(estado)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_orders_nombre_no_asociado ON orders(nombre_no_asociado)")

        # Detalle de órdenes
        db.execute("CREATE INDEX IF NOT EXISTS idx_order_items_order_id ON order_items(order_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_order_items_product_id ON order_items(product_id)")

        # Movimientos de stock
        db.execute("CREATE INDEX IF NOT EXISTS idx_stock_movements_product_id ON stock_movements(product_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_stock_movements_order_id ON stock_movements(order_id)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_stock_movements_fecha ON stock_movements(fecha)")

        # Auditoría
        db.execute("CREATE INDEX IF NOT EXISTS idx_audit_logs_fecha ON audit_logs(fecha)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_audit_logs_accion ON audit_logs(accion)")
        db.execute("CREATE INDEX IF NOT EXISTS idx_audit_logs_admin_id ON audit_logs(admin_id)")

        db.commit()

    except sqlite3.OperationalError as e:
        print("Aviso al verificar estructura:", e)

# =====================================================
# FUNCIÓN AUXILIAR PARA VALIDAR IMÁGENES
# =====================================================
def allowed_file(filename):
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]
    )


# =====================================================
# FUNCIÓN AUXILIAR: NORMALIZAR TEXTO
# =====================================================
def norm_text(value):
    value = (value or "").strip().lower()
    value = value.replace("é", "e")
    return value


# =====================================================
# FUNCIÓN AUXILIAR: VALIDAR CONTRASEÑAS
# =====================================================
def validar_contrasena_segura(password):
    errores = []

    if not password:
        errores.append("La contraseña no puede estar vacía.")
        return errores

    if len(password) < 8:
        errores.append("Debe tener al menos 8 caracteres.")

    if not any(c.isupper() for c in password):
        errores.append("Debe incluir al menos una letra mayúscula.")

    if not any(c.islower() for c in password):
        errores.append("Debe incluir al menos una letra minúscula.")

    if not any(c.isdigit() for c in password):
        errores.append("Debe incluir al menos un número.")

    return errores

# =====================================================
# FUNCIONES AUXILIARES: IMPORTACIÓN DE PRODUCTOS EXCEL
# =====================================================
def importar_parse_numero(value, default=0):
    if value is None or str(value).strip() == "":
        return default

    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    text = text.replace("₡", "").replace(" ", "").replace(",", "")

    return float(text)


def importar_parse_entero(value, default=0):
    return int(importar_parse_numero(value, default))


def importar_parse_activo(value):
    if value is None or str(value).strip() == "":
        return 1

    text = str(value).strip().lower()

    if text in ("1", "activo", "si", "sí", "true", "verdadero"):
        return 1

    if text in ("0", "inactivo", "no", "false", "falso"):
        return 0

    raise ValueError("Activo debe ser 1, 0, activo o inactivo.")


def analizar_excel_productos(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    headers = []
    for cell in ws[1]:
        headers.append(str(cell.value).strip().lower() if cell.value else "")

    columnas_requeridas = ["nombre", "precio", "stock_a_sumar", "stock_minimo", "activo"]
    faltantes = [c for c in columnas_requeridas if c not in headers]

    if faltantes:
        raise ValueError(f"Faltan columnas requeridas: {', '.join(faltantes)}")

    idx = {nombre: headers.index(nombre) for nombre in columnas_requeridas}

    db = get_db()

    preview = []
    errores_generales = []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):

        # Omitir filas completamente vacías
        if not row or all((v is None or str(v).strip() == "") for v in row):
            continue

        fila = {
            "fila": row_num,
            "nombre": "",
            "precio": 0,
            "stock_a_sumar": 0,
            "stock_minimo": 0,
            "activo": 1,
            "accion": "",
            "stock_actual": "",
            "stock_resultante": "",
            "estado": "OK",
            "mensaje": ""
        }

        try:
            nombre = str(row[idx["nombre"]] or "").strip()

            if not nombre:
                raise ValueError("Nombre vacío.")

            precio_num = float(importar_parse_numero(row[idx["precio"]], None))
            stock_sumar = importar_parse_entero(row[idx["stock_a_sumar"]], 0)
            stock_minimo_num = importar_parse_entero(row[idx["stock_minimo"]], 0)
            activo = importar_parse_activo(row[idx["activo"]])

            if precio_num < 0:
                raise ValueError("El precio no puede ser negativo.")

            if stock_minimo_num < 0:
                raise ValueError("El stock mínimo no puede ser negativo.")

            producto = db.execute(
                """
                SELECT id, nombre, stock
                FROM products
                WHERE lower(trim(nombre)) = lower(trim(?))
                """,
                (nombre,)
            ).fetchone()

            fila["nombre"] = nombre
            fila["precio"] = precio_num
            fila["stock_a_sumar"] = stock_sumar
            fila["stock_minimo"] = stock_minimo_num
            fila["activo"] = activo

            if producto:
                stock_actual = int(producto["stock"])
                stock_resultante = stock_actual + stock_sumar

                fila["accion"] = "Actualizar"
                fila["stock_actual"] = stock_actual
                fila["stock_resultante"] = stock_resultante

                if stock_resultante < 0:
                    raise ValueError(
                        f"El ajuste dejaría stock negativo. Stock actual: {stock_actual}."
                    )

            else:
                fila["accion"] = "Crear"
                fila["stock_actual"] = "Nuevo"
                fila["stock_resultante"] = stock_sumar

                if stock_sumar < 0:
                    raise ValueError("No se puede crear un producto con stock inicial negativo.")

            fila["mensaje"] = "Listo para importar."

        except Exception as e:
            fila["estado"] = "Error"
            fila["mensaje"] = str(e)

        preview.append(fila)

    if not preview:
        errores_generales.append("El archivo no contiene filas válidas para procesar.")

    return preview, errores_generales

# =====================================================
# FUNCIÓN AUXILIAR: REGISTRAR AUDITORÍA
# =====================================================
def registrar_auditoria(accion, detalle="", commit=False):
    try:
        db = get_db()

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        admin_id = session.get("user_id")
        admin_nombre = session.get("nombre", "Sistema")

        db.execute(
            """
            INSERT INTO audit_logs (fecha, admin_id, admin_nombre, accion, detalle)
            VALUES (?, ?, ?, ?, ?)
            """,
            (fecha, admin_id, admin_nombre, accion, detalle)
        )

        if commit:
            db.commit()

    except Exception as e:
        print("Error registrando auditoría:", e)


# =====================================================
# COMANDO CLI: BACKUP DB
# Comando:
# flask --app app.py backup-db
#
# Crea un respaldo ZIP local en la carpeta /backups
# Incluye:
# 1) Base de datos SQLite
# 2) Imágenes subidas
# 3) Archivo informativo
# =====================================================
@app.cli.command("backup-db")
def backup_db_command():
    backup_folder = os.path.join(os.getcwd(), "backups")
    os.makedirs(backup_folder, exist_ok=True)

    fecha_backup = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    backup_filename = f"respaldo_aserve_{fecha_backup}.zip"
    backup_path = os.path.join(backup_folder, backup_filename)

    db_original_path = app.config["DATABASE"]

    if not os.path.exists(db_original_path):
        print("❌ No se encontró la base de datos para generar el respaldo.")
        print(f"Ruta esperada: {db_original_path}")
        return

    with zipfile.ZipFile(backup_path, "w", zipfile.ZIP_DEFLATED) as zip_file:

        # =====================================================
        # 1) RESPALDAR BASE DE DATOS SQLITE
        # Se usa source.backup(destination) para hacer una copia segura
        # aunque la app esté abierta.
        # =====================================================
        temp_db_path = None
        source = None
        destination = None

        try:
            with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temp_db:
                temp_db_path = temp_db.name

            source = sqlite3.connect(db_original_path)
            destination = sqlite3.connect(temp_db_path)

            with destination:
                source.backup(destination)

            zip_file.write(
                temp_db_path,
                arcname="database/aserve.db"
            )

        finally:
            if source:
                source.close()

            if destination:
                destination.close()

            if temp_db_path and os.path.exists(temp_db_path):
                os.remove(temp_db_path)

        # =====================================================
        # 2) RESPALDAR IMÁGENES SUBIDAS
        # Normalmente están en static/uploads
        # =====================================================
        upload_folder_abs = app.config["UPLOAD_FOLDER"]

        if not os.path.isabs(upload_folder_abs):
            upload_folder_abs = os.path.join(os.getcwd(), upload_folder_abs)

        if os.path.exists(upload_folder_abs):
            for root, dirs, files in os.walk(upload_folder_abs):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, upload_folder_abs)

                    zip_file.write(
                        file_path,
                        arcname=os.path.join("uploads", relative_path)
                    )

        # =====================================================
        # 3) ARCHIVO INFORMATIVO
        # =====================================================
        info = f"""RESPALDO ASERVE
Fecha de respaldo: {fecha_backup}

Contenido:
- database/aserve.db
- uploads/

Origen:
Este respaldo fue generado desde comando CLI con:
flask --app app.py backup-db

Notas:
Este archivo contiene la base de datos SQLite y las imágenes subidas al sistema.

Recomendación:
Guardar este archivo en un lugar seguro.
No compartirlo con personas no autorizadas, ya que puede contener información sensible del sistema.
"""
        zip_file.writestr("LEEME_RESPALDO.txt", info)

    print("✅ Respaldo creado correctamente.")
    print(f"Archivo: {backup_path}")

# =====================================================
# COMANDO CLI: RESTAURAR RESPALDO
# Comando:
# flask --app app.py restore-backup "backups/respaldo_aserve_XXXX.zip"
#
# IMPORTANTE:
# - Detener Flask antes de restaurar.
# - Crea un respaldo de emergencia antes de reemplazar datos.
# - Restaura:
#   1) database/aserve.db
#   2) uploads/
# =====================================================
@app.cli.command("restore-backup")
@click.argument("zip_path")
def restore_backup_command(zip_path):
    # =====================================================
    # 1) VALIDAR ARCHIVO ZIP
    # =====================================================
    if not os.path.exists(zip_path):
        print("❌ El archivo ZIP no existe.")
        print(f"Ruta recibida: {zip_path}")
        return

    if not zipfile.is_zipfile(zip_path):
        print("❌ El archivo indicado no es un ZIP válido.")
        return

    db_path = app.config["DATABASE"]
    db_folder = os.path.dirname(db_path)

    upload_folder_abs = app.config["UPLOAD_FOLDER"]

    if not os.path.isabs(upload_folder_abs):
        upload_folder_abs = os.path.join(os.getcwd(), upload_folder_abs)

    os.makedirs(db_folder, exist_ok=True)
    os.makedirs(upload_folder_abs, exist_ok=True)

    # =====================================================
    # 2) VALIDAR CONTENIDO DEL RESPALDO ANTES DE TOCAR DATOS
    # =====================================================
    with zipfile.ZipFile(zip_path, "r") as zip_ref:
        zip_names = zip_ref.namelist()

        if "database/aserve.db" not in zip_names:
            print("❌ El respaldo no contiene database/aserve.db.")
            print("No se restauró nada.")
            return

    fecha_restore = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    backup_folder = os.path.join(os.getcwd(), "backups")
    os.makedirs(backup_folder, exist_ok=True)

    pre_backup_path = os.path.join(
        backup_folder,
        f"pre_restore_aserve_{fecha_restore}.zip"
    )

    # =====================================================
    # 3) CREAR RESPALDO DE EMERGENCIA
    # =====================================================
    print("🟡 Creando respaldo de emergencia antes de restaurar...")

    with zipfile.ZipFile(pre_backup_path, "w", zipfile.ZIP_DEFLATED) as zip_file:

        # -------------------------------------------------
        # Respaldar DB actual de forma segura
        # -------------------------------------------------
        if os.path.exists(db_path):
            temp_db_path = None
            source = None
            destination = None

            try:
                with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temp_db:
                    temp_db_path = temp_db.name

                source = sqlite3.connect(db_path)
                destination = sqlite3.connect(temp_db_path)

                with destination:
                    source.backup(destination)

                zip_file.write(
                    temp_db_path,
                    arcname="database/aserve.db"
                )

            finally:
                if source:
                    source.close()

                if destination:
                    destination.close()

                if temp_db_path and os.path.exists(temp_db_path):
                    os.remove(temp_db_path)

        # -------------------------------------------------
        # Respaldar uploads actuales
        # -------------------------------------------------
        if os.path.exists(upload_folder_abs):
            for root, dirs, files in os.walk(upload_folder_abs):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, upload_folder_abs)

                    zip_file.write(
                        file_path,
                        arcname=os.path.join("uploads", relative_path)
                    )

        info = f"""RESPALDO DE EMERGENCIA ASERVE
Fecha: {fecha_restore}

Este respaldo fue creado automáticamente antes de restaurar otro respaldo.

Contenido:
- database/aserve.db
- uploads/

Si algo sale mal, este archivo permite recuperar el estado anterior.
"""
        zip_file.writestr("LEEME_PRE_RESTORE.txt", info)

    print(f"✅ Respaldo de emergencia creado: {pre_backup_path}")

    # =====================================================
    # 4) EXTRAER RESPALDO EN CARPETA TEMPORAL DE FORMA SEGURA
    # =====================================================
    with tempfile.TemporaryDirectory() as temp_dir:
        restored_db = os.path.join(temp_dir, "database", "aserve.db")
        restored_uploads = os.path.join(temp_dir, "uploads")

        with zipfile.ZipFile(zip_path, "r") as zip_ref:
            for member in zip_ref.infolist():
                member_name = member.filename.replace("\\", "/")

                # Evitar rutas peligrosas
                if member_name.startswith("/") or ".." in member_name.split("/"):
                    print(f"⚠️ Archivo omitido por ruta insegura: {member.filename}")
                    continue

                # Solo restaurar lo esperado
                if not (
                    member_name == "database/aserve.db"
                    or member_name.startswith("uploads/")
                ):
                    continue

                target_path = os.path.join(temp_dir, member_name)
                os.makedirs(os.path.dirname(target_path), exist_ok=True)

                if not member.is_dir():
                    with zip_ref.open(member) as source_file:
                        with open(target_path, "wb") as target_file:
                            shutil.copyfileobj(source_file, target_file)

        # Validar nuevamente después de extraer
        if not os.path.exists(restored_db):
            print("❌ No se pudo extraer database/aserve.db.")
            print("No se restauró nada.")
            print(f"Tu respaldo de emergencia quedó en: {pre_backup_path}")
            return

        # =====================================================
        # 5) RESTAURAR BASE DE DATOS
        # =====================================================
        print("🟡 Restaurando base de datos...")

        if os.path.exists(db_path):
            os.remove(db_path)

        shutil.copy2(restored_db, db_path)

        print("✅ Base de datos restaurada correctamente.")

        # =====================================================
        # 6) RESTAURAR IMÁGENES
        # =====================================================
        print("🟡 Restaurando imágenes...")

        if os.path.exists(upload_folder_abs):
            shutil.rmtree(upload_folder_abs)

        if os.path.exists(restored_uploads):
            shutil.copytree(restored_uploads, upload_folder_abs)
            print("✅ Imágenes restauradas correctamente.")
        else:
            os.makedirs(upload_folder_abs, exist_ok=True)
            print("⚠️ El respaldo no tenía carpeta uploads. Se creó vacía.")

    # =====================================================
    # 7) FINAL
    # =====================================================
    print("")
    print("✅ RESTAURACIÓN COMPLETADA CORRECTAMENTE")
    print(f"Respaldo restaurado: {zip_path}")
    print(f"Respaldo de emergencia previo: {pre_backup_path}")
    print("")
    print("Ahora podés iniciar Flask nuevamente con:")
    print("python app.py")

# =====================================================
# COMANDOS DE CONSOLA
# =====================================================
@app.cli.command("init-db")
def init_db_command():
    os.makedirs(app.instance_path, exist_ok=True)

    db = get_db()

    with app.open_resource("schema.sql") as f:
        db.executescript(f.read().decode("utf-8"))

    ensure_app_schema(db)

    db.commit()

    print("✅ Base de datos inicializada correctamente.")


@app.cli.command("create-admin")
def create_admin_command():
    db = get_db()
    ensure_app_schema(db)

    nombre = "Administrador"
    usuario = "admin"
    contrasena = "Admin1234"

    existe = db.execute(
        "SELECT id FROM users WHERE usuario = ?",
        (usuario,)
    ).fetchone()

    if existe:
        print("⚠️ El usuario admin ya existe.")
        return

    hash_pw = bcrypt.hashpw(
        contrasena.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")

    try:
        db.execute(
            """
            INSERT INTO users (nombre, usuario, contrasena_hash, rol, estado, password_temporal)
            VALUES (?, ?, ?, ?, ?, 0)
            """,
            (nombre, usuario, hash_pw, "admin", "activo")
        )
    except sqlite3.OperationalError:
        db.execute(
            """
            INSERT INTO users (nombre, usuario, contrasena_hash, rol, estado)
            VALUES (?, ?, ?, ?, ?)
            """,
            (nombre, usuario, hash_pw, "admin", "activo")
        )

    db.commit()

    print("✅ Admin creado correctamente.")
    print("Usuario: admin | Contraseña: Admin1234")


# =====================================================
# RUTA PRINCIPAL
# =====================================================
@app.route("/")
def inicio():
    return redirect(url_for("catalogo"))


# =====================================================
# LOGIN
# =====================================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        contrasena = (request.form.get("contrasena") or "").strip()

        db = get_db()
        ensure_app_schema(db)

        user = db.execute(
            "SELECT * FROM users WHERE usuario = ?",
            (usuario,)
        ).fetchone()

        if not user:
            flash("Usuario o contraseña incorrectos.", "danger")
            return redirect(url_for("login"))

        if user["estado"] == "bloqueado":
            flash("Usuario bloqueado. Contacte al administrador.", "warning")
            return redirect(url_for("login"))

        if not bcrypt.checkpw(
            contrasena.encode("utf-8"),
            user["contrasena_hash"].encode("utf-8")
        ):
            flash("Usuario o contraseña incorrectos.", "danger")
            return redirect(url_for("login"))

        session["user_id"] = user["id"]
        session["nombre"] = user["nombre"]
        session["rol"] = user["rol"]
        session["password_temporal"] = user["password_temporal"] if "password_temporal" in user.keys() else 0

        if session.get("password_temporal") == 1:
            flash("Debes cambiar tu contraseña temporal antes de continuar.", "warning")
            return redirect(url_for("cambiar_password_temporal"))

        flash(f"Bienvenido/a, {user['nombre']}!", "success")

        if user["rol"] == "admin":
            return redirect(url_for("admin_panel"))

        return redirect(url_for("catalogo"))

    return render_template("login.html")


# =====================================================
# LOGOUT
# =====================================================
@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada correctamente.", "info")
    return redirect(url_for("login"))


# =====================================================
# PROTECCIÓN GLOBAL DE ACCESO
# Objetivo:
# - Si no hay sesión, solo se permite ver login y archivos estáticos.
# - Si el usuario tiene contraseña temporal, solo puede cambiarla o cerrar sesión.
# =====================================================
@app.before_request
def proteger_acceso_global():
    # Nombre de la ruta actual
    endpoint = request.endpoint

    # Algunas solicitudes pueden no tener endpoint definido
    if endpoint is None:
        return None

    # =====================================================
    # RUTAS PÚBLICAS PERMITIDAS SIN LOGIN
    # =====================================================
    rutas_publicas = {
        "login",
        "static"
    }

    # Permitir acceso a login y archivos estáticos
    if endpoint in rutas_publicas:
        return None

    # =====================================================
    # SI NO HAY SESIÓN, REDIRIGIR SIEMPRE AL LOGIN
    # =====================================================
    if "user_id" not in session:
        flash("Debes iniciar sesión para acceder a ASERVE.", "warning")
        return redirect(url_for("login"))

    # =====================================================
    # RUTAS PERMITIDAS CON CONTRASEÑA TEMPORAL
    # =====================================================
    rutas_password_temporal = {
        "cambiar_password_temporal",
        "logout",
        "static"
    }

    # Si el usuario tiene contraseña temporal, solo puede cambiarla o salir
    if session.get("password_temporal") == 1:
        if endpoint not in rutas_password_temporal:
            flash("Debes cambiar tu contraseña temporal antes de continuar.", "warning")
            return redirect(url_for("cambiar_password_temporal"))

    return None

# =====================================================
# PERFIL
# =====================================================
@app.route("/perfil", methods=["GET", "POST"])
def perfil():
    if "user_id" not in session:
        flash("Debes iniciar sesión para ver tu perfil.", "warning")
        return redirect(url_for("login"))

    db = get_db()
    user_id = session["user_id"]

    user = db.execute(
        "SELECT id, nombre, usuario, rol, estado FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        session.clear()
        flash("Sesión inválida. Inicia sesión nuevamente.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        actual = (request.form.get("password_actual") or "").strip()
        nueva = (request.form.get("password_nueva") or "").strip()
        confirmar = (request.form.get("password_confirmar") or "").strip()

        if not actual or not nueva or not confirmar:
            flash("Completa todos los campos para cambiar contraseña.", "danger")
            return redirect(url_for("perfil"))

        if nueva != confirmar:
            flash("La nueva contraseña y su confirmación no coinciden.", "danger")
            return redirect(url_for("perfil"))

        errores_password = validar_contrasena_segura(nueva)

        if errores_password:
            flash(" ".join(errores_password), "warning")
            return redirect(url_for("perfil"))

        user_full = db.execute(
            "SELECT contrasena_hash FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()

        if not user_full or not bcrypt.checkpw(
            actual.encode("utf-8"),
            user_full["contrasena_hash"].encode("utf-8")
        ):
            flash("La contraseña actual no es correcta.", "danger")
            return redirect(url_for("perfil"))

        nuevo_hash = bcrypt.hashpw(
            nueva.encode("utf-8"),
            bcrypt.gensalt()
        ).decode("utf-8")

        db.execute(
            "UPDATE users SET contrasena_hash = ? WHERE id = ?",
            (nuevo_hash, user_id)
        )
        db.commit()

        flash("✅ Contraseña actualizada correctamente.", "success")
        return redirect(url_for("perfil"))

    return render_template("perfil.html", user=user)


# =====================================================
# CAMBIO OBLIGATORIO DE CONTRASEÑA TEMPORAL
# =====================================================
@app.route("/cambiar-password-temporal", methods=["GET", "POST"])
def cambiar_password_temporal():
    if "user_id" not in session:
        flash("Debes iniciar sesión.", "warning")
        return redirect(url_for("login"))

    db = get_db()
    ensure_app_schema(db)

    user_id = session["user_id"]

    user = db.execute(
        "SELECT id, nombre, usuario, password_temporal FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        session.clear()
        flash("Sesión inválida. Inicia sesión nuevamente.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        nueva = (request.form.get("password_nueva") or "").strip()
        confirmar = (request.form.get("password_confirmar") or "").strip()

        if not nueva or not confirmar:
            flash("Completa todos los campos.", "danger")
            return redirect(url_for("cambiar_password_temporal"))

        if nueva != confirmar:
            flash("La nueva contraseña y su confirmación no coinciden.", "danger")
            return redirect(url_for("cambiar_password_temporal"))

        errores_password = validar_contrasena_segura(nueva)

        if errores_password:
            flash(" ".join(errores_password), "warning")
            return redirect(url_for("cambiar_password_temporal"))

        nuevo_hash = bcrypt.hashpw(
            nueva.encode("utf-8"),
            bcrypt.gensalt()
        ).decode("utf-8")

        db.execute(
            """
            UPDATE users
            SET contrasena_hash = ?, password_temporal = 0
            WHERE id = ?
            """,
            (nuevo_hash, user_id)
        )
        db.commit()

        session["password_temporal"] = 0

        flash("Contraseña actualizada correctamente. Ya puedes usar el sistema.", "success")

        if session.get("rol") == "admin":
            return redirect(url_for("admin_panel"))

        return redirect(url_for("catalogo"))

    return render_template("cambiar_password_temporal.html", user=user)


# =====================================================
# PANEL ADMINISTRADOR
# =====================================================
@app.route("/admin")
def admin_panel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if session.get("rol") != "admin":
        flash("No tienes permisos.", "danger")
        return redirect(url_for("inicio"))

    db = get_db()

    creditos_pendientes = db.execute("""
        SELECT COUNT(*) AS n
        FROM orders
        WHERE replace(lower(trim(tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(estado)) = 'pendiente'
    """).fetchone()["n"]

    stock_bajo = db.execute("""
        SELECT COUNT(*) AS n
        FROM products
        WHERE activo = 1
          AND stock_minimo > 0
          AND stock <= stock_minimo
    """).fetchone()["n"]

    usuarios_bloqueados = db.execute("""
        SELECT COUNT(*) AS n
        FROM users
        WHERE estado = 'bloqueado'
    """).fetchone()["n"]

    ordenes_14 = db.execute("""
        SELECT COUNT(*) AS n
        FROM orders
        WHERE date(fecha) >= date('now','-13 day')
          AND date(fecha) <= date('now')
    """).fetchone()["n"]

    return render_template(
        "admin.html",
        kpi_creditos=creditos_pendientes,
        kpi_stock_bajo=stock_bajo,
        kpi_bloqueados=usuarios_bloqueados,
        kpi_ordenes_14=ordenes_14
    )

# =====================================================
# ADMIN: AUDITORÍA
# Ruta: /admin/audit
# - Muestra resumen por tipo de acción
# - Filtra automáticamente últimos 14 días
# - Permite filtrar por acción y fechas
# - Permite ver u ocultar el detalle individual
# =====================================================
@app.route("/admin/audit")
def admin_audit_logs():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    # Asegura que la tabla audit_logs exista
    ensure_app_schema(db)

    accion = (request.args.get("accion") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    ver_detalle = (request.args.get("ver_detalle") or "0").strip() == "1"

    # =====================================================
    # DEFAULT: últimos 14 días
    # Si no se envían fechas desde el filtro, se cargan
    # automáticamente desde hace 13 días hasta hoy.
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    # =====================================================
    # ACCIONES DISPONIBLES PARA EL SELECT
    # Se muestran todas las acciones registradas.
    # =====================================================
    acciones_disponibles = db.execute(
        """
        SELECT DISTINCT accion
        FROM audit_logs
        WHERE accion IS NOT NULL
          AND TRIM(accion) <> ''
        ORDER BY accion ASC
        """
    ).fetchall()

    # =====================================================
    # WHERE DINÁMICO PARA FILTROS
    # =====================================================
    where = []
    params = []

    where.append("date(fecha) >= date(?)")
    params.append(start)

    where.append("date(fecha) <= date(?)")
    params.append(end)

    if accion:
        where.append("accion = ?")
        params.append(accion)

    where_sql = "WHERE " + " AND ".join(where)

    # =====================================================
    # RESUMEN GENERAL
    # =====================================================
    resumen_base = db.execute(
        f"""
        SELECT
            COUNT(id) AS total_acciones,
            COUNT(DISTINCT accion) AS tipos_acciones
        FROM audit_logs
        {where_sql}
        """,
        params
    ).fetchone()

    ultima = db.execute(
        f"""
        SELECT
            fecha,
            accion
        FROM audit_logs
        {where_sql}
        ORDER BY fecha DESC, id DESC
        LIMIT 1
        """,
        params
    ).fetchone()

    if ultima:
        ultima_fecha = ultima["fecha"]
        ultima_accion_nombre = ultima["accion"]
    else:
        ultima_fecha = None
        ultima_accion_nombre = None

    resumen_general = {
        "total_acciones": resumen_base["total_acciones"] if resumen_base else 0,
        "tipos_acciones": resumen_base["tipos_acciones"] if resumen_base else 0,
        "ultima_fecha": ultima_fecha,
        "ultima_accion": ultima_accion_nombre
    }

    # =====================================================
    # RESUMEN POR ACCIÓN
    # =====================================================
    resumen_acciones = db.execute(
        f"""
        SELECT
            accion,
            COUNT(id) AS cantidad,
            MAX(fecha) AS ultima_fecha
        FROM audit_logs
        {where_sql}
        GROUP BY accion
        ORDER BY cantidad DESC, ultima_fecha DESC
        """,
        params
    ).fetchall()

    # =====================================================
    # DETALLE INDIVIDUAL
    # Solo se consulta cuando el usuario pide ver detalle
    # o entra desde una acción específica.
    # =====================================================
    logs = []

    if ver_detalle or accion:
        logs = db.execute(
            f"""
            SELECT
                id,
                fecha,
                admin_nombre,
                accion,
                detalle
            FROM audit_logs
            {where_sql}
            ORDER BY fecha DESC, id DESC
            LIMIT 300
            """,
            params
        ).fetchall()

    return render_template(
        "admin_audit_logs.html",
        accion=accion,
        acciones_disponibles=acciones_disponibles,
        start=start,
        end=end,
        ver_detalle=ver_detalle,
        resumen_general=resumen_general,
        resumen_acciones=resumen_acciones,
        logs=logs
    )

# =====================================================
# ADMIN: DESCARGAR RESPALDO
# Ruta: /admin/backup/download
# - Descarga un ZIP con:
#   1) Base de datos SQLite
#   2) Imágenes subidas de productos
#   3) Archivo informativo
# =====================================================
@app.route("/admin/backup/download")
def admin_backup_download():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db_original_path = app.config["DATABASE"]

    # Validar que exista la base de datos
    if not os.path.exists(db_original_path):
        flash("No se encontró la base de datos para generar el respaldo.", "danger")
        return redirect(url_for("admin_panel"))

    # Asegurar estructura adicional antes de registrar auditoría
    db = get_db()
    ensure_app_schema(db)

    fecha_backup = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"respaldo_aserve_{fecha_backup}.zip"

    zip_buffer = BytesIO()

    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:

        # =====================================================
        # 1) RESPALDAR BASE DE DATOS SQLITE
        # Se usa source.backup(destination) para hacer una copia segura
        # aunque la app esté abierta.
        # =====================================================
        temp_db_path = None
        source = None
        destination = None

        try:
            with tempfile.NamedTemporaryFile(suffix=".db", delete=False) as temp_db:
                temp_db_path = temp_db.name

            source = sqlite3.connect(db_original_path)
            destination = sqlite3.connect(temp_db_path)

            with destination:
                source.backup(destination)

            zip_file.write(
                temp_db_path,
                arcname="database/aserve.db"
            )

        finally:
            if source:
                source.close()

            if destination:
                destination.close()

            if temp_db_path and os.path.exists(temp_db_path):
                os.remove(temp_db_path)

        # =====================================================
        # 2) RESPALDAR IMÁGENES SUBIDAS
        # Normalmente están en static/uploads
        # =====================================================
        upload_folder_abs = app.config["UPLOAD_FOLDER"]

        if not os.path.isabs(upload_folder_abs):
            upload_folder_abs = os.path.join(os.getcwd(), upload_folder_abs)

        if os.path.exists(upload_folder_abs):
            for root, dirs, files in os.walk(upload_folder_abs):
                for file in files:
                    file_path = os.path.join(root, file)
                    relative_path = os.path.relpath(file_path, upload_folder_abs)

                    zip_file.write(
                        file_path,
                        arcname=os.path.join("uploads", relative_path)
                    )

        # =====================================================
        # 3) ARCHIVO INFORMATIVO
        # =====================================================
        info = f"""RESPALDO ASERVE
Fecha de respaldo: {fecha_backup}

Contenido:
- database/aserve.db
- uploads/

Notas:
Este archivo contiene la base de datos SQLite y las imágenes subidas al sistema.

Recomendación:
Guardar este archivo en un lugar seguro.
No compartirlo con personas no autorizadas, ya que puede contener información sensible del sistema.
"""

        zip_file.writestr("LEEME_RESPALDO.txt", info)

    zip_buffer.seek(0)

    # =====================================================
    # AUDITORÍA
    # =====================================================
    registrar_auditoria(
        "Descargar respaldo",
        f"Se generó y descargó el respaldo '{filename}'. Incluye base de datos e imágenes del sistema.",
        commit=True
    )

    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype="application/zip"
    )

# =====================================================
# ADMIN: USUARIOS
# =====================================================
@app.route("/admin/users")
def admin_users():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    usuarios = db.execute("""
        SELECT id, nombre, usuario, rol, estado
        FROM users
        ORDER BY nombre
    """).fetchall()

    return render_template("admin_users.html", usuarios=usuarios)


@app.route("/admin/users/add", methods=["GET", "POST"])
def admin_user_add():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()
    ensure_app_schema(db)

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        usuario = (request.form.get("usuario") or "").strip()
        rol = (request.form.get("rol") or "").strip()
        password = (request.form.get("password") or "").strip()

        if not nombre or not usuario or not password:
            flash("Nombre, usuario y contraseña son obligatorios.", "danger")
            return redirect(url_for("admin_user_add"))

        errores_password = validar_contrasena_segura(password)

        if errores_password:
            flash(" ".join(errores_password), "warning")
            return redirect(url_for("admin_user_add"))

        if rol not in ("admin", "asociado"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("admin_user_add"))

        existe = db.execute(
            "SELECT id FROM users WHERE usuario = ?",
            (usuario,)
        ).fetchone()

        if existe:
            flash("Ese usuario ya existe.", "warning")
            return redirect(url_for("admin_user_add"))

        hash_pw = bcrypt.hashpw(
            password.encode("utf-8"),
            bcrypt.gensalt()
        ).decode("utf-8")

        db.execute(
            """
            INSERT INTO users (nombre, usuario, contrasena_hash, rol, estado, password_temporal)
            VALUES (?, ?, ?, ?, 'activo', 1)
            """,
            (nombre, usuario, hash_pw, rol)
        )

        registrar_auditoria(
            "Crear usuario",
            f"Nombre: {nombre}. Usuario: {usuario}. Rol asignado: {rol}. Estado inicial: activo. Contraseña marcada como temporal para cambio obligatorio."
        )

        db.commit()

        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_add.html")


@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
def admin_user_edit(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()
    ensure_app_schema(db)

    u = db.execute(
        "SELECT id, nombre, usuario, rol, estado FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not u:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        usuario = (request.form.get("usuario") or "").strip()
        rol = (request.form.get("rol") or "").strip()
        estado = (request.form.get("estado") or "").strip()
        new_password = (request.form.get("new_password") or "").strip()

        if not nombre or not usuario:
            flash("Nombre y usuario son obligatorios.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        if rol not in ("admin", "asociado"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        if estado not in ("activo", "bloqueado"):
            flash("Estado inválido.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        existe = db.execute(
            "SELECT id FROM users WHERE usuario = ? AND id <> ?",
            (usuario, user_id)
        ).fetchone()

        if existe:
            flash("Ese usuario ya está en uso por otra cuenta.", "warning")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        db.execute(
            """
            UPDATE users
            SET nombre = ?, usuario = ?, rol = ?, estado = ?
            WHERE id = ?
            """,
            (nombre, usuario, rol, estado, user_id)
        )

        if new_password:
            errores_password = validar_contrasena_segura(new_password)

            if errores_password:
                flash(" ".join(errores_password), "warning")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            hash_pw = bcrypt.hashpw(
                new_password.encode("utf-8"),
                bcrypt.gensalt()
            ).decode("utf-8")

            db.execute(
                """
                UPDATE users
                SET contrasena_hash = ?, password_temporal = 1
                WHERE id = ?
                """,
                (hash_pw, user_id)
            )

            registrar_auditoria(
                "Reset contraseña",
                f"Usuario ID {user_id}. Nombre: {nombre}. Usuario: {usuario}. Se actualizó la contraseña y quedó marcada como temporal para cambio obligatorio en el próximo inicio de sesión."
            )

        registrar_auditoria(
            "Editar usuario",
            f"Usuario ID {user_id}. Nombre: {nombre}. Usuario: {usuario}. Rol: {rol}. Estado: {estado}."
        )

        db.commit()

        flash("Usuario actualizado correctamente.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_edit.html", u=u)


@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
def admin_user_toggle(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if session.get("user_id") == user_id:
        flash("No podés bloquear tu propia cuenta.", "warning")
        return redirect(url_for("admin_users"))

    db = get_db()

    u = db.execute(
        "SELECT id, nombre, usuario, estado FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not u:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))

    nuevo_estado = "activo" if u["estado"] == "bloqueado" else "bloqueado"

    db.execute(
        "UPDATE users SET estado = ? WHERE id = ?",
        (nuevo_estado, user_id)
    )

    registrar_auditoria(
        "Cambiar estado usuario",
        f"Usuario: {u['nombre']}. Código/usuario: {u['usuario']}. Nuevo estado: {nuevo_estado}."
    )

    db.commit()

    flash(f"Estado actualizado a: {nuevo_estado}", "success")
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>/delete", methods=["POST"])
def admin_user_delete(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if session.get("user_id") == user_id:
        flash("No podés eliminar tu propia cuenta.", "warning")
        return redirect(url_for("admin_users"))

    db = get_db()

    user = db.execute(
        "SELECT id, nombre, usuario, rol FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))

    compras = db.execute(
        "SELECT COUNT(*) AS total FROM orders WHERE user_id = ?",
        (user_id,)
    ).fetchone()["total"]

    if compras > 0:
        flash(
            "No se puede eliminar este usuario porque tiene compras registradas. "
            "Podés bloquearlo para impedir el acceso y conservar el historial.",
            "warning"
        )
        return redirect(url_for("admin_users"))

    if user["rol"] == "admin":
        admins = db.execute(
            "SELECT COUNT(*) AS total FROM users WHERE rol = 'admin'"
        ).fetchone()["total"]

        if admins <= 1:
            flash("No se puede eliminar el último administrador del sistema.", "warning")
            return redirect(url_for("admin_users"))

    db.execute(
        "DELETE FROM users WHERE id = ?",
        (user_id,)
    )

    registrar_auditoria(
        "Eliminar usuario",
        f"Nombre: {user['nombre']}. Usuario: {user['usuario']}. Rol: {user['rol']}. Eliminado porque no tenía compras registradas."
    )

    db.commit()

    flash(f"Usuario '{user['nombre']}' eliminado correctamente.", "success")
    return redirect(url_for("admin_users"))

# =====================================================
# MIS COMPRAS
# - Historial del usuario logueado
# - Filtra por tipo de pago y fechas
# - Si no se envían fechas, muestra automáticamente
#   los últimos 14 días incluyendo el día actual
# =====================================================
@app.route("/mis-compras")
def mis_compras():
    if "user_id" not in session:
        flash("Debes iniciar sesión para ver tu historial.", "warning")
        return redirect(url_for("login"))

    user_id = session["user_id"]
    db = get_db()

    pago = (request.args.get("pago") or "").strip().lower()
    pago = pago.replace("é", "e")

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    # =====================================================
    # DEFAULT: últimos 14 días
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = ["o.user_id = ?"]
    params = [user_id]

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    total_rango = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "mis_compras.html",
        orders=orders,
        pago=pago,
        start=start,
        end=end,
        total_rango=total_rango
    )

# =====================================================
# MIS COMPRAS: DETALLE DE COMPRA DEL USUARIO
# - Solo permite ver órdenes del usuario logueado
# - Evita que un usuario vea compras de otro usuario
# =====================================================
@app.route("/mis-compras/<int:order_id>")
def mis_compras_detalle(order_id):
    if "user_id" not in session:
        flash("Debes iniciar sesión.", "warning")
        return redirect(url_for("login"))

    user_id = session["user_id"]
    db = get_db()

    orden = db.execute(
        """
        SELECT
            id,
            fecha,
            tipo_pago,
            total
        FROM orders
        WHERE id = ?
          AND user_id = ?
        """,
        (order_id, user_id)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada o sin permisos para verla.", "danger")
        return redirect(url_for("mis_compras"))

    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        ORDER BY p.nombre ASC
        """,
        (order_id,)
    ).fetchall()

    return render_template(
        "mis_compras_detalle.html",
        orden=orden,
        items=items
    )

# =====================================================
# ADMIN: HISTORIAL DE ÓRDENES
# - Muestra historial general de compras
# - Filtra por estado, tipo de pago, comprador/código y fechas
# - Regla de búsqueda:
#   830  = busca código/usuario
#   #830 = busca número exacto de orden
#   Ana  = busca nombre de comprador
# =====================================================
@app.route("/admin/orders")
def admin_orders():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    estado = (request.args.get("estado") or "").strip().lower()
    pago = (request.args.get("pago") or "").strip().lower().replace("é", "e")
    q = (request.args.get("q") or "").strip()

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    # =====================================================
    # DEFAULT: últimos 14 días
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    # =====================================================
    # FILTRO POR ESTADO
    # =====================================================
    if estado in ("pagada", "pendiente"):
        where.append("lower(trim(o.estado)) = ?")
        params.append(estado)
    else:
        estado = ""

    # =====================================================
    # FILTRO POR TIPO DE PAGO
    # =====================================================
    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)
    else:
        pago = ""

    # =====================================================
    # BÚSQUEDA
    # Reglas:
    # - #15 busca exactamente la orden 15
    # - 830 busca código/usuario 830
    # - Ana busca por nombre de comprador
    # =====================================================
    if q:
        q_clean = q.strip()

        if q_clean.startswith("#"):
            orden_busqueda = q_clean.replace("#", "", 1).strip()

            if orden_busqueda.isdigit():
                where.append("o.id = ?")
                params.append(int(orden_busqueda))
            else:
                # Si escriben algo como #abc, no devuelve resultados
                where.append("1 = 0")

        elif q_clean.isdigit():
            where.append(
                """
                (
                    u.usuario = ?
                    OR u.nombre LIKE ?
                    OR o.nombre_no_asociado LIKE ?
                )
                """
            )
            params.extend([
                q_clean,
                f"%{q_clean}%",
                f"%{q_clean}%"
            ])

        else:
            where.append(
                """
                (
                    u.nombre LIKE ?
                    OR u.usuario LIKE ?
                    OR o.nombre_no_asociado LIKE ?
                )
                """
            )
            params.extend([
                f"%{q_clean}%",
                f"%{q_clean}%",
                f"%{q_clean}%"
            ])

    # =====================================================
    # FILTRO POR FECHAS
    # =====================================================
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    # =====================================================
    # RESUMEN DEL FILTRO
    # =====================================================
    resumen = db.execute(
        f"""
        SELECT
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_filtro
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        """,
        params
    ).fetchone()

    # =====================================================
    # LISTA DE ÓRDENES
    # =====================================================
    orders = db.execute(
        f"""
        SELECT
            o.*,
            u.nombre AS nombre_usuario,
            u.usuario AS codigo_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    return render_template(
        "admin_orders.html",
        orders=orders,
        estado=estado,
        pago=pago,
        q=q,
        start=start,
        end=end,
        num_ordenes=resumen["num_ordenes"] if resumen else 0,
        total_filtro=float(resumen["total_filtro"]) if resumen else 0
    )

# =====================================================
# ADMIN: EXPORTAR HISTORIAL DE ÓRDENES
# - Respeta los mismos filtros de /admin/orders
# - Regla de búsqueda:
#   830  = busca código/usuario
#   #830 = busca número exacto de orden
#   Ana  = busca nombre de comprador
# - Genera Excel con:
#   1) Órdenes
#   2) Detalle productos
# =====================================================
@app.route("/admin/orders/export")
def admin_orders_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    estado = (request.args.get("estado") or "").strip().lower()
    pago = (request.args.get("pago") or "").strip().lower().replace("é", "e")
    q = (request.args.get("q") or "").strip()

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    # =====================================================
    # DEFAULT: últimos 14 días
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    # =====================================================
    # FILTRO POR ESTADO
    # =====================================================
    if estado in ("pagada", "pendiente"):
        where.append("lower(trim(o.estado)) = ?")
        params.append(estado)
    else:
        estado = ""

    # =====================================================
    # FILTRO POR PAGO
    # =====================================================
    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)
    else:
        pago = ""

    # =====================================================
    # BÚSQUEDA
    # Reglas:
    # - #15 busca exactamente la orden 15
    # - 830 busca código/usuario 830
    # - Ana busca por nombre de comprador
    # =====================================================
    if q:
        q_clean = q.strip()

        if q_clean.startswith("#"):
            orden_busqueda = q_clean.replace("#", "", 1).strip()

            if orden_busqueda.isdigit():
                where.append("o.id = ?")
                params.append(int(orden_busqueda))
            else:
                where.append("1 = 0")

        elif q_clean.isdigit():
            where.append(
                """
                (
                    u.usuario = ?
                    OR u.nombre LIKE ?
                    OR o.nombre_no_asociado LIKE ?
                )
                """
            )
            params.extend([
                q_clean,
                f"%{q_clean}%",
                f"%{q_clean}%"
            ])

        else:
            where.append(
                """
                (
                    u.nombre LIKE ?
                    OR u.usuario LIKE ?
                    OR o.nombre_no_asociado LIKE ?
                )
                """
            )
            params.extend([
                f"%{q_clean}%",
                f"%{q_clean}%",
                f"%{q_clean}%"
            ])

    # =====================================================
    # FILTRO POR FECHAS
    # =====================================================
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    # =====================================================
    # RESUMEN
    # =====================================================
    resumen = db.execute(
        f"""
        SELECT
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        """,
        params
    ).fetchone()

    # =====================================================
    # HOJA 1: ÓRDENES
    # =====================================================
    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            COALESCE(u.usuario, '') AS codigo_usuario,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    # =====================================================
    # HOJA 2: DETALLE PRODUCTOS
    # =====================================================
    detalle = db.execute(
        f"""
        SELECT
            o.id AS order_id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            COALESCE(u.usuario, '') AS codigo_usuario,
            o.tipo_pago,
            o.estado,
            o.total AS total_orden,
            p.nombre AS producto,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        JOIN order_items oi ON oi.order_id = o.id
        JOIN products p ON p.id = oi.product_id
        {where_sql}
        ORDER BY o.id DESC, p.nombre ASC
        """,
        params
    ).fetchall()

    # =====================================================
    # CREAR EXCEL
    # =====================================================
    wb = Workbook()

    # =====================================================
    # ESTILOS
    # =====================================================
    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")

    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    # =====================================================
    # HOJA 1: ÓRDENES
    # =====================================================
    ws = wb.active
    ws.title = "Órdenes"

    ws.merge_cells("A1:G1")
    ws["A1"] = "Historial de compras - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    # Filtros aplicados
    ws["A3"] = "Desde"
    ws["B3"] = start

    ws["C3"] = "Hasta"
    ws["D3"] = end

    ws["E3"] = "Pago"
    ws["F3"] = pago if pago else "Todos"

    ws["A4"] = "Estado"
    ws["B4"] = estado if estado else "Todos"

    ws["C4"] = "Búsqueda"
    ws["D4"] = q if q else "Sin búsqueda"

    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=7):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_center

    # Resumen
    ws["A6"] = "Resumen"
    ws["A6"].font = font_bold
    ws["A6"].fill = fill_title

    ws["A7"] = "Órdenes"
    ws["B7"] = resumen["num_ordenes"] if resumen else 0

    ws["A8"] = "Total general"
    ws["B8"] = float(resumen["total_general"]) if resumen else 0

    ws["A9"] = "Contado"
    ws["B9"] = float(resumen["total_contado"]) if resumen else 0

    ws["A10"] = "Crédito"
    ws["B10"] = float(resumen["total_credito"]) if resumen else 0

    for row in ws.iter_rows(min_row=6, max_row=10, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row_num in range(7, 11):
        ws.cell(row=row_num, column=1).font = font_bold

    ws["B8"].number_format = '"₡"#,##0'
    ws["B9"].number_format = '"₡"#,##0'
    ws["B10"].number_format = '"₡"#,##0'

    # Tabla de órdenes
    start_row = 12

    headers = [
        "Orden",
        "Fecha",
        "Comprador",
        "Código",
        "Pago",
        "Estado",
        "Total"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for o in orders:
        ws.cell(row=current_row, column=1).value = f"#{o['id']}"
        ws.cell(row=current_row, column=2).value = o["fecha"]
        ws.cell(row=current_row, column=3).value = o["comprador"]
        ws.cell(row=current_row, column=4).value = o["codigo_usuario"] or "No aplica"
        ws.cell(row=current_row, column=5).value = o["tipo_pago"]
        ws.cell(row=current_row, column=6).value = o["estado"]
        ws.cell(row=current_row, column=7).value = float(o["total"])

        for col in range(1, 8):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=7).number_format = '"₡"#,##0'

        current_row += 1

    if not orders:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
        ws.cell(row=current_row, column=1).value = "Sin datos para los filtros seleccionados."
        ws.cell(row=current_row, column=1).alignment = alignment_center

    widths_ordenes = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 16,
        "E": 15,
        "F": 15,
        "G": 16,
    }

    for col_letter, width in widths_ordenes.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A13"
    ws.auto_filter.ref = f"A{start_row}:G{max(start_row, current_row - 1)}"

    # =====================================================
    # HOJA 2: DETALLE PRODUCTOS
    # =====================================================
    ws2 = wb.create_sheet("Detalle productos")

    ws2.merge_cells("A1:K1")
    ws2["A1"] = "Detalle de productos por orden - ASERVE"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_title
    ws2["A1"].alignment = alignment_left

    headers_detalle = [
        "Orden",
        "Fecha",
        "Comprador",
        "Código",
        "Pago",
        "Estado",
        "Total orden",
        "Producto",
        "Cantidad",
        "Precio unitario",
        "Subtotal"
    ]

    start_row_detalle = 3

    for col_num, header in enumerate(headers_detalle, start=1):
        cell = ws2.cell(row=start_row_detalle, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row_detalle + 1

    for d in detalle:
        ws2.cell(row=current_row, column=1).value = f"#{d['order_id']}"
        ws2.cell(row=current_row, column=2).value = d["fecha"]
        ws2.cell(row=current_row, column=3).value = d["comprador"]
        ws2.cell(row=current_row, column=4).value = d["codigo_usuario"] or "No aplica"
        ws2.cell(row=current_row, column=5).value = d["tipo_pago"]
        ws2.cell(row=current_row, column=6).value = d["estado"]
        ws2.cell(row=current_row, column=7).value = float(d["total_orden"])
        ws2.cell(row=current_row, column=8).value = d["producto"]
        ws2.cell(row=current_row, column=9).value = int(d["cantidad"])
        ws2.cell(row=current_row, column=10).value = float(d["precio_unitario"])
        ws2.cell(row=current_row, column=11).value = float(d["subtotal"])

        for col in range(1, 12):
            cell = ws2.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws2.cell(row=current_row, column=7).number_format = '"₡"#,##0'
        ws2.cell(row=current_row, column=10).number_format = '"₡"#,##0'
        ws2.cell(row=current_row, column=11).number_format = '"₡"#,##0'

        current_row += 1

    if not detalle:
        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=11)
        ws2.cell(row=current_row, column=1).value = "Sin detalle de productos para los filtros seleccionados."
        ws2.cell(row=current_row, column=1).alignment = alignment_center

    widths_detalle = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 16,
        "E": 15,
        "F": 15,
        "G": 16,
        "H": 35,
        "I": 12,
        "J": 18,
        "K": 16,
    }

    for col_letter, width in widths_detalle.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A4"
    ws2.auto_filter.ref = f"A{start_row_detalle}:K{max(start_row_detalle, current_row - 1)}"

    # =====================================================
    # DESCARGA
    # =====================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"historial_compras_{start}_a_{end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: DETALLE DE ORDEN
# - Muestra el detalle completo de una orden
# - Permite regresar correctamente según la pantalla de origen:
#   1) Historial general
#   2) Créditos por colaborador
#   3) Historial por comprador
# =====================================================
@app.route("/admin/orders/<int:order_id>")
def admin_order_detail(order_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    # =====================================================
    # CONSULTA DE LA ORDEN
    # También trae nombre y código del usuario si es registrado
    # =====================================================
    orden = db.execute(
        """
        SELECT
            o.*,
            u.nombre AS nombre_usuario,
            u.usuario AS codigo_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.id = ?
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada.", "danger")
        return redirect(url_for("admin_orders"))

    # =====================================================
    # DETALLE DE PRODUCTOS DE LA ORDEN
    # =====================================================
    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        """,
        (order_id,)
    ).fetchall()

    # =====================================================
    # DATOS PARA REGRESAR A LA PANTALLA CORRECTA
    # from_page puede ser:
    # - orders
    # - credits
    # - buyer
    # - reports
    # =====================================================
    from_page = (request.args.get("from_page") or "orders").strip()

    if from_page not in ("orders", "credits", "buyer", "reports"):
        from_page = "orders"

    # Para volver desde créditos
    back_user_id = (request.args.get("user_id") or "").strip()

    # Para volver desde historial por comprador o reportes
    buyer_key = (request.args.get("buyer_key") or "").strip()
    back_start = (request.args.get("start") or "").strip()
    back_end = (request.args.get("end") or "").strip()

    return render_template(
        "admin_order_detail.html",
        orden=orden,
        items=items,
        from_page=from_page,
        back_user_id=back_user_id,
        buyer_key=buyer_key,
        back_start=back_start,
        back_end=back_end
    )
    
# =====================================================
# ADMIN: COMPRADORES
# =====================================================
@app.route("/admin/buyers")
def admin_buyers():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    registrados = db.execute(
        """
        SELECT
            ('u:' || u.id) AS buyer_key,
            u.nombre AS comprador,
            'registrado' AS tipo,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_monto
        FROM orders o
        JOIN users u ON u.id = o.user_id
        GROUP BY u.id, u.nombre
        """
    ).fetchall()

    no_asociados = db.execute(
        """
        SELECT
            ('na:' || o.nombre_no_asociado) AS buyer_key,
            o.nombre_no_asociado AS comprador,
            'no_asociado' AS tipo,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_monto
        FROM orders o
        WHERE o.user_id IS NULL
          AND o.nombre_no_asociado IS NOT NULL
          AND TRIM(o.nombre_no_asociado) <> ''
        GROUP BY o.nombre_no_asociado
        """
    ).fetchall()

    buyers = list(registrados) + list(no_asociados)
    buyers.sort(key=lambda x: (-float(x["total_monto"]), x["comprador"]))

    return render_template("admin_buyers.html", buyers=buyers)


# =====================================================
# ADMIN: HISTORIAL POR COMPRADOR
# Ruta: /admin/buyer
# - Muestra historial de compras de un comprador específico
# - Filtra automáticamente últimos 14 días
# - Permite filtrar por fechas y tipo de pago
# =====================================================
@app.route("/admin/buyer")
def admin_buyer_history():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    key = (request.args.get("key") or "").strip()
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    pago = norm_text(request.args.get("pago"))

    if not key:
        flash("Comprador inválido.", "danger")
        return redirect(url_for("admin_buyers"))

    # =====================================================
    # DEFAULT: últimos 14 días
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    db = get_db()

    where = []
    params = []

    comprador_nombre = ""
    comprador_tipo = ""
    comprador_codigo = ""

    # =====================================================
    # COMPRADOR REGISTRADO
    # key ejemplo: u:3
    # =====================================================
    if key.startswith("u:"):
        comprador_tipo = "registrado"
        user_id = key.split(":", 1)[1]

        where.append("o.user_id = ?")
        params.append(user_id)

        u = db.execute(
            """
            SELECT nombre, usuario
            FROM users
            WHERE id = ?
            """,
            (user_id,)
        ).fetchone()

        if u:
            comprador_nombre = u["nombre"]
            comprador_codigo = u["usuario"]
        else:
            comprador_nombre = "Usuario"
            comprador_codigo = ""

    # =====================================================
    # COMPRADOR NO ASOCIADO
    # key ejemplo: na:Juan Pérez
    # =====================================================
    elif key.startswith("na:"):
        comprador_tipo = "no_asociado"
        nombre = key.split(":", 1)[1]

        where.append("o.user_id IS NULL")
        where.append("o.nombre_no_asociado = ?")
        params.append(nombre)

        comprador_nombre = nombre
        comprador_codigo = ""

    else:
        flash("Comprador inválido.", "danger")
        return redirect(url_for("admin_buyers"))

    # =====================================================
    # FILTRO POR TIPO DE PAGO
    # =====================================================
    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    # =====================================================
    # FILTRO POR FECHAS
    # =====================================================
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    # =====================================================
    # CONSULTA DE ÓRDENES DEL COMPRADOR
    # =====================================================
    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    total_rango = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "admin_buyer_history.html",
        key=key,
        comprador_nombre=comprador_nombre,
        comprador_tipo=comprador_tipo,
        comprador_codigo=comprador_codigo,
        orders=orders,
        total_rango=total_rango,
        start=start,
        end=end,
        pago=pago
    )

# =====================================================
# ADMIN: EXPORTAR HISTORIAL POR COMPRADOR A EXCEL
# - Exporta todas las órdenes de un comprador
# - Filtra por fechas
# - Incluye detalle de productos por orden
# - Si es comprador registrado, incluye código/usuario
# =====================================================
@app.route("/admin/buyer/export")
def admin_buyer_history_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    key = request.args.get("key", "").strip()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()
    pago = norm_text(request.args.get("pago"))

    if not key:
        flash("Comprador inválido para exportar.", "danger")
        return redirect(url_for("admin_buyers"))

    # =====================================================
    # DEFAULT: últimos 14 días
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    db = get_db()

    where = []
    params = []

    comprador_nombre = ""
    comprador_tipo = ""
    comprador_codigo = ""

    # =====================================================
    # COMPRADOR REGISTRADO
    # =====================================================
    if key.startswith("u:"):
        comprador_tipo = "registrado"
        user_id = key.split(":", 1)[1]

        where.append("o.user_id = ?")
        params.append(user_id)

        u = db.execute(
            """
            SELECT nombre, usuario
            FROM users
            WHERE id = ?
            """,
            (user_id,)
        ).fetchone()

        if u:
            comprador_nombre = u["nombre"]
            comprador_codigo = u["usuario"]
        else:
            comprador_nombre = "Usuario"
            comprador_codigo = ""

    # =====================================================
    # COMPRADOR NO ASOCIADO
    # =====================================================
    elif key.startswith("na:"):
        comprador_tipo = "no asociado"
        nombre = key.split(":", 1)[1]

        where.append("o.user_id IS NULL")
        where.append("o.nombre_no_asociado = ?")
        params.append(nombre)

        comprador_nombre = nombre
        comprador_codigo = ""

    else:
        flash("Comprador inválido para exportar.", "danger")
        return redirect(url_for("admin_buyers"))
    # =====================================================
    # FILTRO POR TIPO DE PAGO
    # =====================================================
    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    # =====================================================
    # FILTRO DE FECHAS
    # =====================================================
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    # =====================================================
    # RESUMEN
    # =====================================================
    resumen = db.execute(
        f"""
        SELECT
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_acumulado
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    # =====================================================
    # DETALLE DE PRODUCTOS
    # =====================================================
    rows = db.execute(
        f"""
        SELECT
            o.id AS order_id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total AS total_orden,
            COALESCE(p.nombre, 'Sin producto') AS producto,
            COALESCE(oi.cantidad, 0) AS cantidad,
            COALESCE(oi.precio_unitario, 0) AS precio_unitario,
            COALESCE((oi.cantidad * oi.precio_unitario), 0) AS subtotal
        FROM orders o
        LEFT JOIN order_items oi ON oi.order_id = o.id
        LEFT JOIN products p ON p.id = oi.product_id
        {where_sql}
        ORDER BY o.fecha DESC, o.id DESC, p.nombre ASC
        """,
        params
    ).fetchall()

    # =====================================================
    # CREAR EXCEL
    # =====================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial comprador"

    # Estilos
    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")

    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    # =====================================================
    # ENCABEZADO DEL REPORTE
    # =====================================================
    ws.merge_cells("A1:J1")
    ws["A1"] = "Historial de compras por comprador - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    ws["A3"] = "Comprador"
    ws["B3"] = comprador_nombre

    ws["A4"] = "Tipo"
    ws["B4"] = comprador_tipo

    ws["A5"] = "Código"
    ws["B5"] = comprador_codigo if comprador_codigo else "No aplica"

    ws["A6"] = "Desde"
    ws["B6"] = start

    ws["A7"] = "Hasta"
    ws["B7"] = end
    ws["A8"] = "Tipo de pago"
    ws["B8"] = "Todos" if not pago else ("Crédito" if pago == "credito" else "Contado")

    ws["A9"] = "Resumen"
    ws["A9"].font = font_bold
    ws["A9"].fill = fill_title

    ws["A10"] = "Órdenes"
    ws["B10"] = resumen["num_ordenes"]

    ws["A11"] = "Total acumulado"
    ws["B11"] = float(resumen["total_acumulado"])
    ws["B11"].number_format = '"₡"#,##0'

    for row in ws.iter_rows(min_row=3, max_row=11, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row in range(3, 12):
        ws.cell(row=row, column=1).font = font_bold

    # =====================================================
    # TABLA DE DETALLE
    # =====================================================
    start_row = 13

    headers = [
        "Orden",
        "Fecha",
        "Comprador",
        "Código",
        "Tipo de pago",
        "Estado",
        "Total orden",
        "Producto",
        "Cantidad",
        "Precio unitario",
        "Subtotal"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for r in rows:
        ws.cell(row=current_row, column=1).value = f"#{r['order_id']}"
        ws.cell(row=current_row, column=2).value = r["fecha"]
        ws.cell(row=current_row, column=3).value = comprador_nombre
        ws.cell(row=current_row, column=4).value = comprador_codigo
        ws.cell(row=current_row, column=5).value = r["tipo_pago"]
        ws.cell(row=current_row, column=6).value = r["estado"]
        ws.cell(row=current_row, column=7).value = float(r["total_orden"])
        ws.cell(row=current_row, column=8).value = r["producto"]
        ws.cell(row=current_row, column=9).value = int(r["cantidad"])
        ws.cell(row=current_row, column=10).value = float(r["precio_unitario"])
        ws.cell(row=current_row, column=11).value = float(r["subtotal"])

        for col in range(1, 12):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=7).number_format = '"₡"#,##0'
        ws.cell(row=current_row, column=10).number_format = '"₡"#,##0'
        ws.cell(row=current_row, column=11).number_format = '"₡"#,##0'

        current_row += 1

    if not rows:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=11
        )
        ws.cell(row=current_row, column=1).value = "Sin datos para este comprador en el rango seleccionado."
        ws.cell(row=current_row, column=1).alignment = alignment_center
        ws.cell(row=current_row, column=1).border = thin_border

    # =====================================================
    # AJUSTES VISUALES
    # =====================================================
    widths = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 16,
        "E": 16,
        "F": 14,
        "G": 16,
        "H": 35,
        "I": 12,
        "J": 18,
        "K": 16,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A14"
    ws.auto_filter.ref = f"A{start_row}:K{max(start_row, current_row - 1)}"

    # =====================================================
    # DESCARGAR ARCHIVO
    # =====================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = secure_filename(comprador_nombre) or "comprador"
    filename = f"historial_comprador_{safe_name}_{start}_a_{end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: CRÉDITOS
# =====================================================
@app.route("/admin/credits")
def admin_credits():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    creditos = db.execute(
        """
        SELECT
            u.id AS user_id,
            u.nombre AS comprador,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS deuda_total
        FROM orders o
        JOIN users u ON u.id = o.user_id
        WHERE
            replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito'
            AND lower(trim(o.estado)) = 'pendiente'
        GROUP BY u.id, u.nombre
        ORDER BY deuda_total DESC, u.nombre ASC
        """
    ).fetchall()

    return render_template("admin_credits.html", creditos=creditos)


@app.route("/admin/credits/user/<int:user_id>")
def admin_credits_user(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    user = db.execute(
        "SELECT id, nombre FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        flash("Colaborador no encontrado.", "danger")
        return redirect(url_for("admin_credits"))

    orders = db.execute(
        """
        SELECT id, fecha, total
        FROM orders
        WHERE user_id = ?
          AND replace(lower(trim(tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(estado)) = 'pendiente'
        ORDER BY fecha ASC
        """,
        (user_id,)
    ).fetchall()

    deuda_total = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "admin_credits_user.html",
        user=user,
        orders=orders,
        deuda_total=deuda_total
    )


@app.route("/admin/credits/user/<int:user_id>/pay-bulk", methods=["POST"])
def admin_credits_pay_bulk(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    order_ids = request.form.getlist("order_ids")

    if not order_ids:
        flash("No seleccionaste ninguna orden.", "warning")
        return redirect(url_for("admin_credits_user", user_id=user_id))

    try:
        order_ids_int = [int(x) for x in order_ids]
    except ValueError:
        flash("Selección inválida.", "danger")
        return redirect(url_for("admin_credits_user", user_id=user_id))

    db = get_db()
    placeholders = ",".join(["?"] * len(order_ids_int))

    cur = db.execute(
        f"""
        UPDATE orders
        SET estado = 'pagada'
        WHERE user_id = ?
          AND replace(lower(trim(tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(estado)) = 'pendiente'
          AND id IN ({placeholders})
        """,
        [user_id] + order_ids_int
    )

    comprador = db.execute(
        "SELECT nombre FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    comprador_nombre = comprador["nombre"] if comprador else f"Usuario ID {user_id}"
    ordenes_txt = ", ".join([f"#{oid}" for oid in order_ids_int])

    registrar_auditoria(
        "Pagar créditos en bloque",
        f"Comprador: {comprador_nombre}. Órdenes pagadas: {ordenes_txt}. Total de órdenes actualizadas: {cur.rowcount}."
    )

    db.commit()

    flash(f"Se marcaron como pagadas {cur.rowcount} orden(es).", "success")
    return redirect(url_for("admin_credits_user", user_id=user_id))


@app.route("/admin/credits/pay/<int:order_id>", methods=["POST"])
def admin_credits_pay(order_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    orden = db.execute(
        """
        SELECT
            o.id,
            o.user_id,
            o.total,
            u.nombre AS comprador
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.id = ?
          AND replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(o.estado)) = 'pendiente'
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Esa orden no existe o ya fue pagada.", "warning")
        return redirect(url_for("admin_credits"))

    db.execute(
        "UPDATE orders SET estado = 'pagada' WHERE id = ?",
        (order_id,)
    )

    comprador_nombre = orden["comprador"] if orden["comprador"] else f"Usuario ID {orden['user_id']}"

    registrar_auditoria(
        "Pagar crédito",
        f"Orden #{order_id} marcada como pagada. Comprador: {comprador_nombre}. Monto: ₡{float(orden['total']):,.2f}."
    )

    db.commit()

    flash(f"Orden #{order_id} marcada como pagada.", "success")
    return redirect(url_for("admin_credits_user", user_id=orden["user_id"]))


# =====================================================
# ADMIN: STOCK Y PRODUCTOS
# =====================================================
@app.route("/admin/stock")
def admin_stock():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    only_low = (request.args.get("only_low") or "").strip() == "1"

    db = get_db()

    productos = db.execute(
        """
        SELECT id, nombre, stock, stock_minimo, activo
        FROM products
        ORDER BY nombre ASC
        """
    ).fetchall()

    bajos = []

    for p in productos:
        if int(p["activo"]) == 1 and int(p["stock_minimo"]) > 0 and int(p["stock"]) <= int(p["stock_minimo"]):
            bajos.append(p)

    if only_low:
        productos = bajos

    return render_template(
        "admin_stock.html",
        productos=productos,
        bajos_count=len(bajos),
        only_low=only_low
    )


@app.route("/admin/products")
def admin_products():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    productos = db.execute("""
        SELECT id, nombre, precio, stock, stock_minimo, activo, image_filename
        FROM products
        ORDER BY nombre ASC
    """).fetchall()

    bajos = [
        p for p in productos
        if int(p["activo"]) == 1
        and int(p["stock_minimo"] or 0) > 0
        and int(p["stock"] or 0) <= int(p["stock_minimo"] or 0)
    ]

    return render_template("admin_products.html", productos=productos, bajos=bajos)


@app.route("/admin/products/add", methods=["GET", "POST"])
def admin_product_add():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        precio = (request.form.get("precio") or "").strip()
        stock = (request.form.get("stock") or "").strip()
        stock_minimo = (request.form.get("stock_minimo") or "0").strip()
        activo = 1 if request.form.get("activo") == "on" else 0
        imagen = request.files.get("imagen")

        if not nombre or not precio or not stock:
            flash("Nombre, precio y stock son obligatorios.", "danger")
            return redirect(url_for("admin_product_add"))

        try:
            precio_num = float(precio)
            stock_num = int(stock)
            stock_minimo_num = int(stock_minimo)

            if precio_num < 0 or stock_num < 0 or stock_minimo_num < 0:
                flash("Precio, stock y stock mínimo no pueden ser negativos.", "danger")
                return redirect(url_for("admin_product_add"))

        except ValueError:
            flash("Precio debe ser número y stock/stock mínimo deben ser enteros.", "danger")
            return redirect(url_for("admin_product_add"))

        db = get_db()

        # Evitar productos duplicados por nombre
        existe = db.execute(
            """
            SELECT id
            FROM products
            WHERE lower(trim(nombre)) = lower(trim(?))
            """,
            (nombre,)
        ).fetchone()

        if existe:
            flash("Ya existe un producto con ese nombre. Revisá la gestión de productos antes de crearlo nuevamente.", "warning")
            return redirect(url_for("admin_product_add"))

        image_filename = None

        if imagen and imagen.filename:
            if not allowed_file(imagen.filename):
                flash("Formato de imagen no permitido. Usá PNG, JPG, JPEG o WEBP.", "danger")
                return redirect(url_for("admin_product_add"))

            filename_img = secure_filename(imagen.filename)
            unique_name = f"{int(datetime.now().timestamp())}_{filename_img}"

            upload_folder_abs = app.config["UPLOAD_FOLDER"]
            os.makedirs(upload_folder_abs, exist_ok=True)

            imagen.save(os.path.join(upload_folder_abs, unique_name))
            image_filename = unique_name

        db.execute(
            """
            INSERT INTO products (nombre, precio, stock, stock_minimo, activo, image_filename)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            (nombre, precio_num, stock_num, stock_minimo_num, activo, image_filename)
        )

        registrar_auditoria(
            "Crear producto",
            f"Producto: {nombre}. Precio: ₡{precio_num:,.2f}. Stock inicial: {stock_num}. Stock mínimo: {stock_minimo_num}. Estado: {'activo' if activo == 1 else 'inactivo'}."
        )

        db.commit()

        flash("✅ Producto agregado correctamente.", "success")
        return redirect(url_for("admin_products"))

    return render_template("admin_product_add.html")

@app.route("/admin/products/<int:product_id>/edit", methods=["GET", "POST"])
def admin_product_edit(product_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    p = db.execute(
        """
        SELECT id, nombre, precio, stock, stock_minimo, activo, image_filename
        FROM products
        WHERE id = ?
        """,
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_products"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        precio = (request.form.get("precio") or "").strip()
        stock_minimo = (request.form.get("stock_minimo") or "0").strip()
        activo = 1 if request.form.get("activo") == "on" else 0
        imagen = request.files.get("imagen")

        if not nombre or not precio:
            flash("Nombre y precio son obligatorios.", "danger")
            return redirect(url_for("admin_product_edit", product_id=product_id))

        try:
            precio_num = float(precio)
            stock_minimo_num = int(stock_minimo)

            if precio_num < 0 or stock_minimo_num < 0:
                flash("Precio y stock mínimo no pueden ser negativos.", "danger")
                return redirect(url_for("admin_product_edit", product_id=product_id))

        except ValueError:
            flash("Precio debe ser número y stock mínimo debe ser entero.", "danger")
            return redirect(url_for("admin_product_edit", product_id=product_id))

        # Evitar duplicar nombre con otro producto
        existe = db.execute(
            """
            SELECT id
            FROM products
            WHERE lower(trim(nombre)) = lower(trim(?))
              AND id <> ?
            """,
            (nombre, product_id)
        ).fetchone()

        if existe:
            flash("Ya existe otro producto con ese nombre. Usá un nombre diferente.", "warning")
            return redirect(url_for("admin_product_edit", product_id=product_id))

        image_filename = p["image_filename"]

        if imagen and imagen.filename:
            if not allowed_file(imagen.filename):
                flash("Formato de imagen no permitido. Usá PNG, JPG, JPEG o WEBP.", "danger")
                return redirect(url_for("admin_product_edit", product_id=product_id))

            filename_img = secure_filename(imagen.filename)
            unique_name = f"{int(datetime.now().timestamp())}_{filename_img}"

            upload_folder_abs = app.config["UPLOAD_FOLDER"]
            os.makedirs(upload_folder_abs, exist_ok=True)

            imagen.save(os.path.join(upload_folder_abs, unique_name))
            image_filename = unique_name

        db.execute(
            """
            UPDATE products
            SET nombre = ?, precio = ?, stock_minimo = ?, activo = ?, image_filename = ?
            WHERE id = ?
            """,
            (nombre, precio_num, stock_minimo_num, activo, image_filename, product_id)
        )

        registrar_auditoria(
            "Editar producto",
            f"Producto ID {product_id}. Nombre: {nombre}. Precio: ₡{precio_num:,.2f}. "
            f"Stock actual conservado: {p['stock']}. Stock mínimo: {stock_minimo_num}. "
            f"Estado: {'activo' if activo == 1 else 'inactivo'}."
        )

        db.commit()

        flash("✅ Producto actualizado correctamente.", "success")
        return redirect(url_for("admin_products"))

    return render_template("admin_product_edit.html", p=p)
# =====================================================
# ADMIN: AJUSTAR STOCK DE PRODUCTO
# - Permite sumar o restar stock sin editar manualmente el total
# - Registra movimiento en stock_movements
# - Registra auditoría administrativa
# =====================================================
@app.route("/admin/products/<int:product_id>/stock-adjust", methods=["POST"])
def admin_product_stock_adjust(product_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    producto = db.execute(
        """
        SELECT id, nombre, stock
        FROM products
        WHERE id = ?
        """,
        (product_id,)
    ).fetchone()

    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_products"))

    ajuste_raw = (request.form.get("ajuste_stock") or "").strip()
    motivo = (request.form.get("motivo_stock") or "").strip()

    if not ajuste_raw:
        flash("Debes indicar la cantidad a sumar o restar.", "warning")
        return redirect(url_for("admin_product_edit", product_id=product_id))

    try:
        ajuste = int(ajuste_raw)
    except ValueError:
        flash("La cantidad de ajuste debe ser un número entero.", "danger")
        return redirect(url_for("admin_product_edit", product_id=product_id))

    if ajuste == 0:
        flash("La cantidad de ajuste no puede ser cero.", "warning")
        return redirect(url_for("admin_product_edit", product_id=product_id))

    stock_actual = int(producto["stock"])
    nuevo_stock = stock_actual + ajuste

    if nuevo_stock < 0:
        flash(
            f"No se puede aplicar el ajuste. El stock actual es {stock_actual} y el resultado quedaría negativo.",
            "danger"
        )
        return redirect(url_for("admin_product_edit", product_id=product_id))

    if not motivo:
        motivo = "Ajuste manual de stock"

    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    db.execute(
        """
        UPDATE products
        SET stock = ?
        WHERE id = ?
        """,
        (nuevo_stock, product_id)
    )

    db.execute(
        """
        INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
        VALUES (?, ?, ?, ?, NULL)
        """,
        (product_id, ajuste, motivo, fecha)
    )

    signo = "+" if ajuste > 0 else ""

    registrar_auditoria(
        "Ajuste de stock",
        f"Producto: {producto['nombre']}. Stock anterior: {stock_actual}. Ajuste aplicado: {signo}{ajuste}. Nuevo stock: {nuevo_stock}. Motivo: {motivo}."
    )

    db.commit()

    flash(
        f"Stock actualizado correctamente. Antes: {stock_actual} | Ajuste: {signo}{ajuste} | Nuevo stock: {nuevo_stock}.",
        "success"
    )

    return redirect(url_for("admin_product_edit", product_id=product_id))

# =====================================================
# ADMIN: DESCARGAR PLANTILLA DE IMPORTACIÓN DE PRODUCTOS
# Ruta: /admin/products/import/template
# - Genera un Excel con las columnas requeridas
# =====================================================
@app.route("/admin/products/import/template")
def admin_products_import_template():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla productos"

    headers = ["nombre", "precio", "stock_a_sumar", "stock_minimo", "activo"]
    ws.append(headers)

    # Ejemplos
    ws.append(["Coca Cola 600ml", 800, 10, 3, 1])
    ws.append(["Doritos", 950, 5, 2, 1])
    ws.append(["Producto inactivo ejemplo", 500, 0, 1, 0])

    # Anchos
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12

    # Formatos
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    ws["C1"].font = Font(bold=True)
    ws["D1"].font = Font(bold=True)
    ws["E1"].font = Font(bold=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name="plantilla_importacion_productos_aserve.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: IMPORTAR PRODUCTOS DESDE EXCEL - PREVISUALIZACIÓN
# Ruta: /admin/products/import
# Formato esperado:
# nombre | precio | stock_a_sumar | stock_minimo | activo
# =====================================================
@app.route("/admin/products/import", methods=["GET", "POST"])
def admin_products_import():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if request.method == "GET":
        return render_template("admin_product_import.html")

    archivo = request.files.get("archivo")

    if not archivo or not archivo.filename:
        flash("Debes seleccionar un archivo Excel.", "warning")
        return redirect(url_for("admin_products_import"))

    if not archivo.filename.lower().endswith(".xlsx"):
        flash("El archivo debe ser formato .xlsx.", "danger")
        return redirect(url_for("admin_products_import"))

    # Carpeta temporal para guardar el archivo mientras se confirma
    import_folder = os.path.join(app.instance_path, "imports")
    os.makedirs(import_folder, exist_ok=True)

    import_token = secrets.token_urlsafe(16)
    import_filename = f"productos_import_{import_token}.xlsx"
    import_path = os.path.join(import_folder, import_filename)

    archivo.save(import_path)

    try:
        preview, errores_generales = analizar_excel_productos(import_path)

        total_crear = sum(1 for r in preview if r["estado"] == "OK" and r["accion"] == "Crear")
        total_actualizar = sum(1 for r in preview if r["estado"] == "OK" and r["accion"] == "Actualizar")
        total_errores = sum(1 for r in preview if r["estado"] == "Error") + len(errores_generales)

        return render_template(
            "admin_product_import.html",
            preview=True,
            import_token=import_token,
            preview_rows=preview,
            errores_generales=errores_generales,
            total_crear=total_crear,
            total_actualizar=total_actualizar,
            total_errores=total_errores
        )

    except Exception as e:
        if os.path.exists(import_path):
            os.remove(import_path)

        flash(f"No se pudo leer el archivo Excel: {str(e)}", "danger")
        return redirect(url_for("admin_products_import"))
    # =====================================================
    # FUNCIONES INTERNAS PARA VALIDAR DATOS DEL EXCEL
    # =====================================================
    def parse_numero(value, default=0):
        if value is None or str(value).strip() == "":
            return default

        if isinstance(value, (int, float)):
            return value

        text = str(value).strip()
        text = text.replace("₡", "").replace(" ", "").replace(",", "")

        return float(text)

    def parse_entero(value, default=0):
        return int(parse_numero(value, default))

    def parse_activo(value):
        if value is None or str(value).strip() == "":
            return 1

        text = str(value).strip().lower()

        if text in ("1", "activo", "si", "sí", "true", "verdadero"):
            return 1

        if text in ("0", "inactivo", "no", "false", "falso"):
            return 0

        raise ValueError("Activo debe ser 1, 0, activo o inactivo.")

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        # Leer encabezados
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value).strip().lower() if cell.value else "")

        columnas_requeridas = ["nombre", "precio", "stock_a_sumar", "stock_minimo", "activo"]

        faltantes = [c for c in columnas_requeridas if c not in headers]

        if faltantes:
            flash(f"Faltan columnas requeridas en el Excel: {', '.join(faltantes)}", "danger")
            return redirect(url_for("admin_products_import"))

        idx = {nombre: headers.index(nombre) for nombre in columnas_requeridas}

        db = get_db()

        creados = 0
        actualizados = 0
        stock_ajustado = 0
        errores = []

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                nombre = str(row[idx["nombre"]] or "").strip()

                if not nombre:
                    errores.append(f"Fila {row_num}: nombre vacío.")
                    continue

                precio_num = float(parse_numero(row[idx["precio"]], None))
                stock_sumar = parse_entero(row[idx["stock_a_sumar"]], 0)
                stock_minimo_num = parse_entero(row[idx["stock_minimo"]], 0)
                activo = parse_activo(row[idx["activo"]])

                if precio_num < 0:
                    errores.append(f"Fila {row_num}: el precio no puede ser negativo.")
                    continue

                if stock_minimo_num < 0:
                    errores.append(f"Fila {row_num}: el stock mínimo no puede ser negativo.")
                    continue

                producto = db.execute(
                    """
                    SELECT id, nombre, stock
                    FROM products
                    WHERE lower(trim(nombre)) = lower(trim(?))
                    """,
                    (nombre,)
                ).fetchone()

                # =====================================================
                # PRODUCTO EXISTENTE: actualizar datos y sumar stock
                # =====================================================
                if producto:
                    stock_actual = int(producto["stock"])
                    nuevo_stock = stock_actual + stock_sumar

                    if nuevo_stock < 0:
                        errores.append(
                            f"Fila {row_num}: el ajuste dejaría stock negativo para '{nombre}'. Stock actual: {stock_actual}."
                        )
                        continue

                    db.execute(
                        """
                        UPDATE products
                        SET precio = ?, stock_minimo = ?, activo = ?
                        WHERE id = ?
                        """,
                        (precio_num, stock_minimo_num, activo, producto["id"])
                    )

                    if stock_sumar != 0:
                        db.execute(
                            """
                            UPDATE products
                            SET stock = ?
                            WHERE id = ?
                            """,
                            (nuevo_stock, producto["id"])
                        )

                        db.execute(
                            """
                            INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
                            VALUES (?, ?, ?, ?, NULL)
                            """,
                            (
                                producto["id"],
                                stock_sumar,
                                "Importación Excel de productos",
                                fecha
                            )
                        )

                        stock_ajustado += 1

                    actualizados += 1

                # =====================================================
                # PRODUCTO NUEVO: crear
                # =====================================================
                else:
                    if stock_sumar < 0:
                        errores.append(
                            f"Fila {row_num}: no se puede crear '{nombre}' con stock inicial negativo."
                        )
                        continue

                    cur = db.execute(
                        """
                        INSERT INTO products (nombre, precio, stock, stock_minimo, activo, image_filename)
                        VALUES (?, ?, ?, ?, ?, NULL)
                        """,
                        (nombre, precio_num, stock_sumar, stock_minimo_num, activo)
                    )

                    product_id = cur.lastrowid

                    if stock_sumar != 0:
                        db.execute(
                            """
                            INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
                            VALUES (?, ?, ?, ?, NULL)
                            """,
                            (
                                product_id,
                                stock_sumar,
                                "Importación Excel de productos",
                                fecha
                            )
                        )

                        stock_ajustado += 1

                    creados += 1

            except Exception as e:
                errores.append(f"Fila {row_num}: {str(e)}")

        registrar_auditoria(
            "Importar productos Excel",
            f"Importación masiva completada. Creados: {creados}. Actualizados: {actualizados}. Ajustes de stock: {stock_ajustado}. Errores: {len(errores)}."
        )

        db.commit()

        return render_template(
            "admin_product_import.html",
            resultado=True,
            creados=creados,
            actualizados=actualizados,
            stock_ajustado=stock_ajustado,
            errores=errores
        )

    except Exception as e:
        flash(f"No se pudo procesar el archivo: {str(e)}", "danger")
        return redirect(url_for("admin_products_import"))

# =====================================================
# ADMIN: CONFIRMAR IMPORTACIÓN DE PRODUCTOS
# Ruta: /admin/products/import/confirm
# - Aplica los cambios reales después de la previsualización
# =====================================================
@app.route("/admin/products/import/confirm", methods=["POST"])
def admin_products_import_confirm():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    import_token = (request.form.get("import_token") or "").strip()

    if not import_token:
        flash("No se encontró el archivo de importación para confirmar.", "danger")
        return redirect(url_for("admin_products_import"))

    import_folder = os.path.join(app.instance_path, "imports")
    import_path = os.path.join(import_folder, f"productos_import_{import_token}.xlsx")

    if not os.path.exists(import_path):
        flash("El archivo de importación ya no existe o expiró.", "danger")
        return redirect(url_for("admin_products_import"))

    try:
        preview, errores_generales = analizar_excel_productos(import_path)

        if errores_generales:
            flash("No se puede confirmar porque el archivo tiene errores generales.", "danger")
            return redirect(url_for("admin_products_import"))

        db = get_db()

        creados = 0
        actualizados = 0
        stock_ajustado = 0
        errores = []

        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for fila in preview:
            if fila["estado"] != "OK":
                errores.append(f"Fila {fila['fila']}: {fila['mensaje']}")
                continue

            nombre = fila["nombre"]
            precio_num = float(fila["precio"])
            stock_sumar = int(fila["stock_a_sumar"])
            stock_minimo_num = int(fila["stock_minimo"])
            activo = int(fila["activo"])

            producto = db.execute(
                """
                SELECT id, nombre, stock
                FROM products
                WHERE lower(trim(nombre)) = lower(trim(?))
                """,
                (nombre,)
            ).fetchone()

            if producto:
                stock_actual = int(producto["stock"])
                nuevo_stock = stock_actual + stock_sumar

                if nuevo_stock < 0:
                    errores.append(
                        f"Fila {fila['fila']}: el ajuste dejaría stock negativo para '{nombre}'."
                    )
                    continue

                db.execute(
                    """
                    UPDATE products
                    SET precio = ?, stock_minimo = ?, activo = ?
                    WHERE id = ?
                    """,
                    (precio_num, stock_minimo_num, activo, producto["id"])
                )

                if stock_sumar != 0:
                    db.execute(
                        """
                        UPDATE products
                        SET stock = ?
                        WHERE id = ?
                        """,
                        (nuevo_stock, producto["id"])
                    )

                    db.execute(
                        """
                        INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
                        VALUES (?, ?, ?, ?, NULL)
                        """,
                        (
                            producto["id"],
                            stock_sumar,
                            "Importación Excel de productos",
                            fecha
                        )
                    )

                    stock_ajustado += 1

                actualizados += 1

            else:
                if stock_sumar < 0:
                    errores.append(
                        f"Fila {fila['fila']}: no se puede crear '{nombre}' con stock inicial negativo."
                    )
                    continue

                cur = db.execute(
                    """
                    INSERT INTO products (nombre, precio, stock, stock_minimo, activo, image_filename)
                    VALUES (?, ?, ?, ?, ?, NULL)
                    """,
                    (nombre, precio_num, stock_sumar, stock_minimo_num, activo)
                )

                product_id = cur.lastrowid

                if stock_sumar != 0:
                    db.execute(
                        """
                        INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
                        VALUES (?, ?, ?, ?, NULL)
                        """,
                        (
                            product_id,
                            stock_sumar,
                            "Importación Excel de productos",
                            fecha
                        )
                    )

                    stock_ajustado += 1

                creados += 1

        registrar_auditoria(
            "Importar productos Excel",
            f"Importación confirmada. Creados: {creados}. Actualizados: {actualizados}. Ajustes de stock: {stock_ajustado}. Errores: {len(errores)}."
        )

        db.commit()

        if os.path.exists(import_path):
            os.remove(import_path)

        return render_template(
            "admin_product_import.html",
            resultado=True,
            creados=creados,
            actualizados=actualizados,
            stock_ajustado=stock_ajustado,
            errores=errores
        )

    except Exception as e:
        flash(f"No se pudo confirmar la importación: {str(e)}", "danger")
        return redirect(url_for("admin_products_import"))

# =====================================================
# ADMIN: ELIMINAR IMAGEN DE PRODUCTO
# - Elimina la imagen física del producto
# - Deja image_filename en NULL
# =====================================================
@app.route("/admin/products/<int:product_id>/image/delete", methods=["POST"])
def admin_product_image_delete(product_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    p = db.execute(
        """
        SELECT id, nombre, image_filename
        FROM products
        WHERE id = ?
        """,
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_products"))

    if not p["image_filename"]:
        flash("Este producto no tiene imagen para eliminar.", "warning")
        return redirect(url_for("admin_product_edit", product_id=product_id))

    upload_folder_abs = app.config["UPLOAD_FOLDER"]

    if not os.path.isabs(upload_folder_abs):
        upload_folder_abs = os.path.join(os.getcwd(), upload_folder_abs)

    image_path = os.path.join(
        upload_folder_abs,
        os.path.basename(p["image_filename"])
    )

    if os.path.exists(image_path):
        try:
            os.remove(image_path)
        except OSError as e:
            print("Error eliminando imagen:", e)
            flash("No se pudo eliminar el archivo de imagen.", "danger")
            return redirect(url_for("admin_product_edit", product_id=product_id))

    db.execute(
        """
        UPDATE products
        SET image_filename = NULL
        WHERE id = ?
        """,
        (product_id,)
    )

    registrar_auditoria(
        "Eliminar imagen producto",
        f"Producto ID {product_id}. Producto: {p['nombre']}. Se eliminó la imagen asociada."
    )

    db.commit()

    flash("Imagen eliminada correctamente.", "success")
    return redirect(url_for("admin_product_edit", product_id=product_id))

# =====================================================
# ADMIN: EXPORTAR INVENTARIO A EXCEL
# Ruta: /admin/products/export
# - Descarga un Excel con el inventario actual
# - Incluye productos activos, inactivos, stock, mínimos y valor estimado
# =====================================================
@app.route("/admin/products/export")
def admin_products_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    productos = db.execute(
        """
        SELECT
            id,
            nombre,
            precio,
            stock,
            stock_minimo,
            activo,
            image_filename
        FROM products
        ORDER BY nombre ASC
        """
    ).fetchall()

    # =====================================================
    # RESUMEN GENERAL
    # =====================================================
    total_productos = len(productos)
    total_activos = sum(1 for p in productos if int(p["activo"]) == 1)
    total_inactivos = sum(1 for p in productos if int(p["activo"]) != 1)

    total_bajos = sum(
        1 for p in productos
        if int(p["activo"]) == 1
        and int(p["stock_minimo"] or 0) > 0
        and int(p["stock"] or 0) <= int(p["stock_minimo"] or 0)
    )

    valor_total_inventario = sum(
        float(p["precio"] or 0) * int(p["stock"] or 0)
        for p in productos
    )

    fecha_reporte = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # =====================================================
    # CREAR EXCEL
    # =====================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    # Estilos
    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    # =====================================================
    # TÍTULO
    # =====================================================
    ws.merge_cells("A1:I1")
    ws["A1"] = "Reporte de inventario - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    # =====================================================
    # RESUMEN
    # =====================================================
    ws["A3"] = "Fecha de reporte"
    ws["B3"] = fecha_reporte

    ws["A5"] = "Productos registrados"
    ws["B5"] = total_productos

    ws["C5"] = "Activos"
    ws["D5"] = total_activos

    ws["E5"] = "Inactivos"
    ws["F5"] = total_inactivos

    ws["A6"] = "Stock bajo"
    ws["B6"] = total_bajos

    ws["C6"] = "Valor inventario"
    ws["D6"] = float(valor_total_inventario)

    ws["D6"].number_format = '"₡"#,##0'

    for row in ws.iter_rows(min_row=3, max_row=6, min_col=1, max_col=9):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for cell_ref in ["A3", "A5", "C5", "E5", "A6", "C6"]:
        ws[cell_ref].font = font_bold

    # =====================================================
    # TABLA
    # =====================================================
    start_row = 8

    headers = [
        "ID",
        "Nombre",
        "Precio",
        "Stock actual",
        "Stock mínimo",
        "Estado",
        "Alerta",
        "Valor inventario",
        "Imagen"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for p in productos:
        activo = int(p["activo"])
        stock = int(p["stock"] or 0)
        stock_minimo = int(p["stock_minimo"] or 0)
        precio = float(p["precio"] or 0)

        estado = "Activo" if activo == 1 else "Inactivo"

        if activo == 1 and stock_minimo > 0 and stock <= stock_minimo:
            alerta = "Bajo stock"
        elif activo != 1:
            alerta = "Producto inactivo"
        else:
            alerta = "OK"

        valor_inventario = precio * stock

        ws.cell(row=current_row, column=1).value = p["id"]
        ws.cell(row=current_row, column=2).value = p["nombre"]
        ws.cell(row=current_row, column=3).value = precio
        ws.cell(row=current_row, column=4).value = stock
        ws.cell(row=current_row, column=5).value = stock_minimo
        ws.cell(row=current_row, column=6).value = estado
        ws.cell(row=current_row, column=7).value = alerta
        ws.cell(row=current_row, column=8).value = valor_inventario
        ws.cell(row=current_row, column=9).value = p["image_filename"] or "Sin imagen"

        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=3).number_format = '"₡"#,##0'
        ws.cell(row=current_row, column=8).number_format = '"₡"#,##0'

        current_row += 1

    if not productos:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
        ws.cell(row=current_row, column=1).value = "No hay productos registrados."
        ws.cell(row=current_row, column=1).alignment = alignment_center

    # =====================================================
    # AJUSTES VISUALES
    # =====================================================
    widths = {
        "A": 10,
        "B": 35,
        "C": 14,
        "D": 14,
        "E": 14,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 28,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A9"
    ws.auto_filter.ref = f"A{start_row}:I{max(start_row, current_row - 1)}"

    # =====================================================
    # AUDITORÍA
    # =====================================================
    registrar_auditoria(
        "Exportar inventario",
        f"Se descargó reporte de inventario. Productos: {total_productos}. Activos: {total_activos}. Stock bajo: {total_bajos}.",
        commit=True
    )

    # =====================================================
    # DESCARGA
    # =====================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"inventario_aserve_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: ELIMINAR PRODUCTO
# - Solo elimina productos sin ventas ni movimientos
# - Si tiene historial, se debe desactivar
# =====================================================
@app.route("/admin/products/<int:product_id>/delete", methods=["POST"])
def admin_product_delete(product_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    p = db.execute(
        """
        SELECT id, nombre, image_filename
        FROM products
        WHERE id = ?
        """,
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_products"))

    ventas = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM order_items
        WHERE product_id = ?
        """,
        (product_id,)
    ).fetchone()["total"]

    movimientos = db.execute(
        """
        SELECT COUNT(*) AS total
        FROM stock_movements
        WHERE product_id = ?
        """,
        (product_id,)
    ).fetchone()["total"]

    if ventas > 0 or movimientos > 0:
        flash(
            "No se puede eliminar este producto porque ya tiene ventas o movimientos registrados. "
            "Para conservar el historial, podés dejarlo inactivo.",
            "warning"
        )
        return redirect(url_for("admin_products"))

    # Eliminar imagen física si existe
    if p["image_filename"]:
        upload_folder_abs = app.config["UPLOAD_FOLDER"]

        if not os.path.isabs(upload_folder_abs):
            upload_folder_abs = os.path.join(os.getcwd(), upload_folder_abs)

        image_path = os.path.join(
            upload_folder_abs,
            os.path.basename(p["image_filename"])
        )

        if os.path.exists(image_path):
            try:
                os.remove(image_path)
            except OSError as e:
                print("Error eliminando imagen del producto:", e)

    db.execute(
        "DELETE FROM products WHERE id = ?",
        (product_id,)
    )

    registrar_auditoria(
        "Eliminar producto",
        f"Producto ID {product_id}. Producto: {p['nombre']}. Eliminado porque no tenía ventas ni movimientos registrados."
    )

    db.commit()

    flash(f"Producto '{p['nombre']}' eliminado correctamente.", "success")
    return redirect(url_for("admin_products"))

# =====================================================
# REPORTES
# =====================================================
@app.route("/admin/reports/sales")
def admin_report_sales():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    pago = (request.args.get("pago") or "").strip().lower().replace("é", "e")

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    resumen = db.execute(
        f"""
        SELECT
            COUNT(*) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    return render_template(
        "admin_report_sales.html",
        start=start,
        end=end,
        pago=pago,
        resumen=resumen,
        orders=orders
    )


@app.route("/admin/reports/sales/export")
def admin_report_sales_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()
    pago = (request.args.get("pago") or "").strip().lower().replace("é", "e")

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    resumen = db.execute(
        f"""
        SELECT
            COUNT(*) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte ventas"

    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    ws.merge_cells("A1:F1")
    ws["A1"] = "Reporte de ventas - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    ws["A3"] = "Desde"
    ws["B3"] = start
    ws["C3"] = "Hasta"
    ws["D3"] = end
    ws["E3"] = "Tipo de pago"
    ws["F3"] = pago if pago else "Todos"

    for cell in ws[3]:
        cell.border = thin_border
        cell.alignment = alignment_center
        cell.font = font_bold

    ws["A5"] = "Resumen"
    ws["A5"].font = font_bold
    ws["A5"].fill = fill_title
    ws["A5"].alignment = alignment_left

    ws["A6"] = "Órdenes"
    ws["B6"] = resumen["num_ordenes"]

    ws["A7"] = "Total general"
    ws["B7"] = float(resumen["total_general"])

    ws["A8"] = "Contado"
    ws["B8"] = float(resumen["total_contado"])

    ws["A9"] = "Crédito"
    ws["B9"] = float(resumen["total_credito"])

    for row in ws.iter_rows(min_row=5, max_row=9, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row_num in range(6, 10):
        ws.cell(row=row_num, column=1).font = font_bold

    ws["B7"].number_format = '"₡"#,##0'
    ws["B8"].number_format = '"₡"#,##0'
    ws["B9"].number_format = '"₡"#,##0'

    start_row = 11
    headers = ["Orden", "Fecha", "Comprador", "Pago", "Estado", "Total"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for o in orders:
        ws.cell(row=current_row, column=1).value = f"#{o['id']}"
        ws.cell(row=current_row, column=2).value = o["fecha"]
        ws.cell(row=current_row, column=3).value = o["comprador"]
        ws.cell(row=current_row, column=4).value = o["tipo_pago"]
        ws.cell(row=current_row, column=5).value = o["estado"]
        ws.cell(row=current_row, column=6).value = float(o["total"])

        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=6).number_format = '"₡"#,##0'
        current_row += 1

    if not orders:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1).value = "Sin datos para el filtro seleccionado."
        ws.cell(row=current_row, column=1).alignment = alignment_center

    widths = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 15,
        "E": 15,
        "F": 16,
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A12"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_ventas_{start}_a_{end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/admin/reports/top-products")
def admin_report_top_products():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    top = db.execute(
        f"""
        SELECT
            p.id,
            p.nombre,
            SUM(oi.cantidad) AS unidades,
            COALESCE(SUM(oi.cantidad * oi.precio_unitario), 0) AS monto
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        JOIN products p ON p.id = oi.product_id
        {where_sql}
        GROUP BY p.id, p.nombre
        ORDER BY unidades DESC, monto DESC
        LIMIT 20
        """,
        params
    ).fetchall()

    return render_template(
        "admin_report_top_products.html",
        start=start,
        end=end,
        top=top
    )


# =====================================================
# CATÁLOGO
# - Muestra únicamente productos activos y con stock
# - Permite búsqueda por nombre
# - Evita mostrar productos sin disponibilidad
# =====================================================
@app.route("/catalogo")
def catalogo():
    db = get_db()

    q = (request.args.get("q") or "").strip()

    where = [
        "activo = 1",
        "stock > 0"
    ]

    params = []

    if q:
        where.append("nombre LIKE ?")
        params.append(f"%{q}%")

    where_sql = "WHERE " + " AND ".join(where)

    productos = db.execute(
        f"""
        SELECT
            id,
            nombre,
            precio,
            stock,
            stock_minimo,
            activo,
            image_filename
        FROM products
        {where_sql}
        ORDER BY nombre ASC
        """,
        params
    ).fetchall()

    return render_template(
        "catalogo.html",
        productos=productos,
        q=q
    )

# =====================================================
# CARRITO
# =====================================================
def get_cart():
    if "cart" not in session:
        session["cart"] = {}

    return session["cart"]


# =====================================================
# CARRITO: VER CARRITO
# - Limpia productos que ya no existen
# - Quita productos inactivos
# - Quita productos sin stock
# - Ajusta cantidades si superan el stock actual
# =====================================================
@app.route("/carrito")
def ver_carrito():
    cart = get_cart()

    if not cart:
        return render_template("carrito.html", items=[], total=0)

    db = get_db()

    ids = list(cart.keys())
    placeholders = ",".join(["?"] * len(ids))

    productos = db.execute(
        f"""
        SELECT *
        FROM products
        WHERE id IN ({placeholders})
        """,
        ids
    ).fetchall()

    productos_map = {str(p["id"]): p for p in productos}

    items = []
    total = 0
    cart_limpio = {}
    hubo_ajustes = False

    for pid, cantidad_cart in cart.items():
        p = productos_map.get(str(pid))

        # Si el producto ya no existe en base de datos, se elimina del carrito
        if not p:
            hubo_ajustes = True
            continue

        # Validar cantidad del carrito
        try:
            cantidad = int(cantidad_cart)
        except ValueError:
            hubo_ajustes = True
            continue

        # Si la cantidad es inválida, se elimina del carrito
        if cantidad <= 0:
            hubo_ajustes = True
            continue

        # Si el producto está inactivo, se elimina del carrito
        if int(p["activo"]) != 1:
            hubo_ajustes = True
            continue

        stock_disponible = int(p["stock"])

        # Si el producto no tiene stock, se elimina del carrito
        if stock_disponible <= 0:
            hubo_ajustes = True
            continue

        # Si el carrito tiene más cantidad que el stock disponible, se ajusta
        if cantidad > stock_disponible:
            cantidad = stock_disponible
            hubo_ajustes = True

        subtotal = float(p["precio"]) * cantidad
        total += subtotal

        cart_limpio[str(p["id"])] = cantidad

        items.append({
            "id": p["id"],
            "nombre": p["nombre"],
            "precio": float(p["precio"]),
            "stock": stock_disponible,
            "image_filename": p["image_filename"],
            "cantidad": cantidad,
            "subtotal": subtotal
        })

    session["cart"] = cart_limpio

    if hubo_ajustes:
        flash(
            "El carrito fue actualizado porque algunos productos cambiaron de disponibilidad o stock.",
            "warning"
        )

    return render_template("carrito.html", items=items, total=total)


# =====================================================
# CARRITO: AGREGAR PRODUCTO
# =====================================================
@app.route("/carrito/add/<int:product_id>", methods=["POST"])
def carrito_add(product_id):
    db = get_db()

    p = db.execute(
        """
        SELECT *
        FROM products
        WHERE id = ?
          AND activo = 1
        """,
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no disponible.", "danger")
        return redirect(url_for("catalogo"))

    if int(p["stock"]) <= 0:
        flash("Este producto no tiene stock disponible.", "warning")
        return redirect(url_for("catalogo"))

    cart = get_cart()
    pid = str(product_id)

    try:
        cantidad_actual = int(cart.get(pid, 0))
    except ValueError:
        cantidad_actual = 0

    if cantidad_actual + 1 > int(p["stock"]):
        flash("No hay suficiente stock para agregar más.", "warning")
        return redirect(url_for("catalogo"))

    cart[pid] = cantidad_actual + 1
    session["cart"] = cart

    flash("Producto agregado al carrito.", "success")
    return redirect(url_for("catalogo"))


# =====================================================
# CARRITO: AUMENTAR CANTIDAD
# =====================================================
@app.route("/carrito/inc/<int:product_id>", methods=["POST"])
def carrito_inc(product_id):
    db = get_db()

    p = db.execute(
        """
        SELECT *
        FROM products
        WHERE id = ?
          AND activo = 1
        """,
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no disponible.", "danger")
        return redirect(url_for("ver_carrito"))

    if int(p["stock"]) <= 0:
        flash("Este producto ya no tiene stock disponible.", "warning")
        return redirect(url_for("ver_carrito"))

    cart = get_cart()
    pid = str(product_id)

    try:
        cantidad_actual = int(cart.get(pid, 0))
    except ValueError:
        cantidad_actual = 0

    if cantidad_actual + 1 > int(p["stock"]):
        flash("No hay suficiente stock.", "warning")
        return redirect(url_for("ver_carrito"))

    cart[pid] = cantidad_actual + 1
    session["cart"] = cart

    return redirect(url_for("ver_carrito"))


# =====================================================
# CARRITO: DISMINUIR CANTIDAD
# =====================================================
@app.route("/carrito/dec/<int:product_id>", methods=["POST"])
def carrito_dec(product_id):
    cart = get_cart()
    pid = str(product_id)

    try:
        cantidad_actual = int(cart.get(pid, 0))
    except ValueError:
        cantidad_actual = 0

    if cantidad_actual <= 1:
        cart.pop(pid, None)
    else:
        cart[pid] = cantidad_actual - 1

    session["cart"] = cart

    return redirect(url_for("ver_carrito"))


# =====================================================
# CARRITO: ELIMINAR PRODUCTO
# =====================================================
@app.route("/carrito/remove/<int:product_id>", methods=["POST"])
def carrito_remove(product_id):
    cart = get_cart()
    cart.pop(str(product_id), None)

    session["cart"] = cart

    return redirect(url_for("ver_carrito"))


# =====================================================
# CARRITO: VACIAR CARRITO
# =====================================================
@app.route("/carrito/clear", methods=["POST"])
def carrito_clear():
    session["cart"] = {}
    flash("Carrito vaciado.", "info")

    return redirect(url_for("ver_carrito"))

# =====================================================
# CHECKOUT
# - Solo usuarios con sesión pueden comprar
# - Valida stock antes de confirmar
# - Evita reenvío normal del formulario con token interno
# - Calcula total siempre desde base de datos
# - Usa bloqueo transaccional para proteger inventario
# - Descuenta stock con validación directa en UPDATE
# =====================================================
@app.route("/checkout", methods=["GET", "POST"])
def checkout():
    # =====================================================
    # SEGURIDAD: SOLO USUARIOS CON SESIÓN
    # Aunque ya existe protección global, se valida también aquí.
    # =====================================================
    if "user_id" not in session:
        flash("Debes iniciar sesión para realizar compras en ASERVE.", "warning")
        return redirect(url_for("login"))

    cart = get_cart()

    if not cart:
        flash("Tu carrito está vacío.", "info")
        return redirect(url_for("catalogo"))

    db = get_db()

    # =====================================================
    # VALIDAR CARRITO
    # Convierte IDs y cantidades a valores seguros.
    # =====================================================
    try:
        cart_limpio = {}

        for pid, cantidad in cart.items():
            pid_limpio = str(int(pid))
            cantidad_limpia = int(cantidad)

            if cantidad_limpia <= 0:
                flash("Hay una cantidad inválida en el carrito.", "warning")
                return redirect(url_for("ver_carrito"))

            cart_limpio[pid_limpio] = cantidad_limpia

    except ValueError:
        session["cart"] = {}
        session.pop("checkout_token", None)
        flash("El carrito tenía datos inválidos y fue limpiado.", "warning")
        return redirect(url_for("catalogo"))

    if not cart_limpio:
        session["cart"] = {}
        session.pop("checkout_token", None)
        flash("Tu carrito está vacío.", "info")
        return redirect(url_for("catalogo"))

    # =====================================================
    # DATOS DEL USUARIO LOGUEADO
    # =====================================================
    user_id = session.get("user_id")
    rol = session.get("rol")

    # La tabla orders solo acepta: asociado / no_asociado.
    # Como ahora toda compra requiere sesión, registramos como asociado.
    tipo_usuario = "asociado"

    # Admin y asociado pueden usar crédito.
    can_credit = True if rol in ("admin", "asociado") else False

    # =====================================================
    # GET: MOSTRAR PANTALLA DE CONFIRMACIÓN
    # =====================================================
    if request.method == "GET":
        ids = list(cart_limpio.keys())
        placeholders = ",".join(["?"] * len(ids))

        productos = db.execute(
            f"""
            SELECT *
            FROM products
            WHERE id IN ({placeholders})
            """,
            ids
        ).fetchall()

        productos_map = {str(p["id"]): p for p in productos}

        items = []
        total = 0

        for pid, cantidad in cart_limpio.items():
            p = productos_map.get(pid)

            if not p:
                flash("Uno de los productos del carrito ya no existe.", "warning")
                return redirect(url_for("ver_carrito"))

            if int(p["activo"]) != 1:
                flash(f"El producto '{p['nombre']}' ya no está disponible.", "warning")
                return redirect(url_for("ver_carrito"))

            if cantidad > int(p["stock"]):
                flash(f"No hay stock suficiente de '{p['nombre']}'.", "warning")
                return redirect(url_for("ver_carrito"))

            subtotal = float(p["precio"]) * cantidad
            total += subtotal

            items.append({
                "id": p["id"],
                "nombre": p["nombre"],
                "precio": float(p["precio"]),
                "cantidad": cantidad,
                "subtotal": subtotal
            })

        # Token interno para evitar reenvío normal del formulario
        checkout_token = secrets.token_urlsafe(24)
        session["checkout_token"] = checkout_token
        session.modified = True

        return render_template(
            "checkout.html",
            items=items,
            total=total,
            rol=rol,
            es_no_asociado=False,
            can_credit=can_credit,
            checkout_token=checkout_token
        )

    # =====================================================
    # POST: CONFIRMAR COMPRA
    # =====================================================
    token_form = (request.form.get("checkout_token") or "").strip()
    token_session = session.get("checkout_token")

    if not token_session or token_form != token_session:
        session.pop("checkout_token", None)
        flash("Esta compra ya fue procesada o el formulario expiró. Revisá tu carrito antes de continuar.", "warning")
        return redirect(url_for("ver_carrito"))

    tipo_pago = norm_text(request.form.get("tipo_pago"))

    if tipo_pago not in ("contado", "credito"):
        flash("Tipo de pago inválido.", "danger")
        return redirect(url_for("checkout"))

    # =====================================================
    # VALIDAR PERMISOS DE CRÉDITO
    # =====================================================
    if tipo_pago == "credito" and not can_credit:
        flash("Tu usuario no tiene permiso para comprar a crédito.", "danger")
        return redirect(url_for("checkout"))

    try:
        # =====================================================
        # TRANSACCIÓN
        # BEGIN IMMEDIATE bloquea escritura para proteger stock
        # =====================================================
        db.execute("BEGIN IMMEDIATE")

        ids = list(cart_limpio.keys())
        placeholders = ",".join(["?"] * len(ids))

        productos = db.execute(
            f"""
            SELECT *
            FROM products
            WHERE id IN ({placeholders})
            """,
            ids
        ).fetchall()

        productos_map = {str(p["id"]): p for p in productos}

        items = []
        total = 0

        # =====================================================
        # VALIDAR NUEVAMENTE STOCK DESDE BASE DE DATOS
        # =====================================================
        for pid, cantidad in cart_limpio.items():
            p = productos_map.get(pid)

            if not p:
                db.rollback()
                flash("Uno de los productos del carrito ya no existe.", "warning")
                return redirect(url_for("ver_carrito"))

            if int(p["activo"]) != 1:
                db.rollback()
                flash(f"El producto '{p['nombre']}' ya no está disponible.", "warning")
                return redirect(url_for("ver_carrito"))

            if cantidad > int(p["stock"]):
                db.rollback()
                flash(f"No hay stock suficiente de '{p['nombre']}'. Stock disponible: {p['stock']}.", "warning")
                return redirect(url_for("ver_carrito"))

            subtotal = float(p["precio"]) * cantidad
            total += subtotal

            items.append({
                "id": p["id"],
                "nombre": p["nombre"],
                "precio": float(p["precio"]),
                "cantidad": cantidad,
                "subtotal": subtotal
            })

        estado = "pagada" if tipo_pago == "contado" else "pendiente"
        fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # =====================================================
        # CREAR ORDEN
        # =====================================================
        cur = db.execute(
            """
            INSERT INTO orders (
                fecha,
                tipo_usuario,
                user_id,
                nombre_no_asociado,
                tipo_pago,
                total,
                estado
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                fecha,
                tipo_usuario,
                user_id,
                None,
                tipo_pago,
                float(total),
                estado
            )
        )

        order_id = cur.lastrowid

        # =====================================================
        # INSERTAR DETALLE Y DESCONTAR STOCK
        # =====================================================
        for it in items:
            db.execute(
                """
                INSERT INTO order_items (
                    order_id,
                    product_id,
                    cantidad,
                    precio_unitario
                )
                VALUES (?, ?, ?, ?)
                """,
                (
                    order_id,
                    it["id"],
                    it["cantidad"],
                    float(it["precio"])
                )
            )

            stock_update = db.execute(
                """
                UPDATE products
                SET stock = stock - ?
                WHERE id = ?
                  AND activo = 1
                  AND stock >= ?
                """,
                (
                    it["cantidad"],
                    it["id"],
                    it["cantidad"]
                )
            )

            if stock_update.rowcount != 1:
                db.rollback()
                flash(f"No se pudo descontar stock de '{it['nombre']}'. Intentá nuevamente.", "warning")
                return redirect(url_for("ver_carrito"))

            db.execute(
                """
                INSERT INTO stock_movements (
                    product_id,
                    cambio_stock,
                    motivo,
                    fecha,
                    order_id
                )
                VALUES (?, ?, ?, ?, ?)
                """,
                (
                    it["id"],
                    -it["cantidad"],
                    "Venta",
                    fecha,
                    order_id
                )
            )

        db.commit()

        # =====================================================
        # LIMPIAR CARRITO Y TOKEN
        # =====================================================
        session["cart"] = {}
        session.pop("checkout_token", None)

        session["last_order_id"] = order_id
        session.modified = True

        flash(f"✅ Compra realizada. Orden #{order_id} creada.", "success")
        return redirect(url_for("order_success", order_id=order_id))

    except Exception as e:
        db.rollback()
        print("Error en checkout:", e)
        flash("Ocurrió un error al procesar la compra. Intentá nuevamente.", "danger")
        return redirect(url_for("ver_carrito"))

# =====================================================
# ORDEN SUCCESS
# - Muestra comprobante de compra
# - Protege la orden para que no se pueda ver cambiando el ID en la URL
# - Asociado: solo ve sus propias órdenes
# - Admin: puede ver cualquier orden
# - No se permite acceso sin sesión
# =====================================================
@app.route("/orden/<int:order_id>")
def order_success(order_id):
    # =====================================================
    # SEGURIDAD: SOLO USUARIOS CON SESIÓN
    # =====================================================
    if "user_id" not in session:
        flash("Debes iniciar sesión para ver una orden.", "warning")
        return redirect(url_for("login"))

    db = get_db()

    orden = db.execute(
        """
        SELECT
            o.*,
            u.nombre AS nombre_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.id = ?
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada.", "danger")
        return redirect(url_for("catalogo"))

    # =====================================================
    # SEGURIDAD DE ACCESO A LA ORDEN
    # =====================================================
    usuario_logueado = session.get("user_id")
    rol = session.get("rol")

    puede_ver = False

    # Admin puede ver cualquier comprobante
    if rol == "admin":
        puede_ver = True

    # Asociado solo puede ver sus propias órdenes
    elif orden["user_id"] == usuario_logueado:
        puede_ver = True

    if not puede_ver:
        flash("No tenés permiso para ver esta orden.", "warning")
        return redirect(url_for("mis_compras"))

    # =====================================================
    # NOMBRE DEL COMPRADOR
    # - Si es usuario registrado, usa nombre_usuario.
    # - Si es orden antigua de no asociado, usa nombre_no_asociado.
    # =====================================================
    comprador = orden["nombre_usuario"] if orden["user_id"] else orden["nombre_no_asociado"]

    # =====================================================
    # DETALLE DE PRODUCTOS
    # =====================================================
    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        """,
        (order_id,)
    ).fetchall()

    return render_template(
        "order_success.html",
        orden=orden,
        items=items,
        comprador=comprador
    )

# =====================================================
# EJECUCIÓN LOCAL
# =====================================================
if __name__ == "__main__":
    os.makedirs(app.instance_path, exist_ok=True)

    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"

    app.run(debug=debug_mode)