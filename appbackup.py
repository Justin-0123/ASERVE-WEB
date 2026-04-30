# =====================================================
# IMPORTS GENERALES
# =====================================================
import os
import sqlite3
import bcrypt
from datetime import datetime, timedelta
from io import BytesIO

from flask import (
    Flask, render_template, request,
    redirect, url_for, session, flash, g, send_file
)

from werkzeug.utils import secure_filename

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# =====================================================
# CONFIGURACIÓN PRINCIPAL DE FLASK
# =====================================================
app = Flask(__name__)

# Clave para manejar sesiones (en producción debe ser segura)
app.secret_key = "clave_secreta_aserve"

# Ruta de la base de datos SQLite (se guarda en /instance/aserve.db)
app.config["DATABASE"] = os.path.join(app.instance_path, "aserve.db")

# Configuración para subida de imágenes de productos
app.config["UPLOAD_FOLDER"] = os.path.join("static", "uploads")
app.config["ALLOWED_EXTENSIONS"] = {"png", "jpg", "jpeg", "webp"}


# =====================================================
# FUNCIONES DE BASE DE DATOS
# =====================================================
def get_db():
    """
    Devuelve una conexión a la base de datos.
    Usa 'g' para mantener una sola conexión por request.
    """
    if "db" not in g:
        g.db = sqlite3.connect(app.config["DATABASE"])
        g.db.row_factory = sqlite3.Row
    return g.db


@app.teardown_appcontext
def close_db(exception=None):
    """
    Cierra la conexión a la base de datos al finalizar cada request.
    """
    db = g.pop("db", None)
    if db is not None:
        db.close()


# =====================================================
# FUNCIÓN AUXILIAR PARA VALIDAR IMÁGENES
# =====================================================
def allowed_file(filename):
    """
    Verifica que el archivo tenga una extensión permitida
    para subir como imagen.
    """
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in app.config["ALLOWED_EXTENSIONS"]
    )


# =====================================================
# FUNCIÓN AUXILIAR: NORMALIZAR TEXTO (PAGOS / ESTADOS)
# =====================================================
def norm_text(value):
    value = (value or "").strip().lower()
    value = value.replace("é", "e")
    return value

# =====================================================
# FUNCIÓN AUXILIAR: VALIDAR CONTRASEÑAS
# Reglas:
# - mínimo 8 caracteres
# - al menos una mayúscula
# - al menos una minúscula
# - al menos un número
# =====================================================
def validar_contrasena_segura(password):
    errores = []

    if not password:
        errores.append("La contraseña no puede estar vacía.")
        return errores

    if len(password) < 8:
        errores.append("Debe tener al menos 8 caracteres.")

    if not any(c.isupper() for c in password):
        errores.append("Debe incluir al menos una letra mayúscula.")

    if not any(c.islower() for c in password):
        errores.append("Debe incluir al menos una letra minúscula.")

    if not any(c.isdigit() for c in password):
        errores.append("Debe incluir al menos un número.")

    return errores

# =====================================================
# COMANDOS DE CONSOLA (FLASK CLI)
# =====================================================
@app.cli.command("init-db")
def init_db_command():
    """
    Inicializa la base de datos ejecutando schema.sql
    """
    os.makedirs(app.instance_path, exist_ok=True)
    db = get_db()
    with app.open_resource("schema.sql") as f:
        db.executescript(f.read().decode("utf-8"))
    db.commit()
    print("✅ Base de datos inicializada correctamente.")


@app.cli.command("create-admin")
def create_admin_command():
    """
    Crea el usuario administrador inicial si no existe
    """
    db = get_db()

    nombre = "Administrador"
    usuario = "admin"
    contrasena = "Admin1234"

    existe = db.execute(
        "SELECT id FROM users WHERE usuario = ?",
        (usuario,)
    ).fetchone()

    if existe:
        print("⚠️ El usuario admin ya existe.")
        return

    hash_pw = bcrypt.hashpw(
        contrasena.encode("utf-8"),
        bcrypt.gensalt()
    ).decode("utf-8")

    db.execute(
        "INSERT INTO users (nombre, usuario, contrasena_hash, rol, estado) VALUES (?, ?, ?, ?, ?)",
        (nombre, usuario, hash_pw, "admin", "activo")
    )
    db.commit()

    print("✅ Admin creado correctamente.")
    print("Usuario: admin | Contraseña: Admin1234")


# =====================================================
# RUTA: /
# =====================================================
@app.route("/")
def inicio():
    return redirect(url_for("catalogo"))


# =====================================================
# LOGIN
# =====================================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = (request.form.get("usuario") or "").strip()
        contrasena = (request.form.get("contrasena") or "").strip()

        db = get_db()
        user = db.execute(
            "SELECT * FROM users WHERE usuario = ?",
            (usuario,)
        ).fetchone()

        if not user:
            flash("Usuario o contraseña incorrectos.", "danger")
            return redirect(url_for("login"))

        if user["estado"] == "bloqueado":
            flash("Usuario bloqueado. Contacte al administrador.", "warning")
            return redirect(url_for("login"))

        if not bcrypt.checkpw(
            contrasena.encode("utf-8"),
            user["contrasena_hash"].encode("utf-8")
        ):
            flash("Usuario o contraseña incorrectos.", "danger")
            return redirect(url_for("login"))

        # Guardar datos en sesión
        session["user_id"] = user["id"]
        session["nombre"] = user["nombre"]
        session["rol"] = user["rol"]

        flash(f"Bienvenido/a, {user['nombre']}!", "success")

        # Redirección según rol
        if user["rol"] == "admin":
            return redirect(url_for("admin_panel"))

        return redirect(url_for("catalogo"))

    return render_template("login.html")


# =====================================================
# LOGOUT
# =====================================================
@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada correctamente.", "info")
    return redirect(url_for("login"))


# =====================================================
# PERFIL (ASOCIADO/ADMIN LOGUEADO)
# =====================================================
@app.route("/perfil", methods=["GET", "POST"])
def perfil():
    if "user_id" not in session:
        flash("Debes iniciar sesión para ver tu perfil.", "warning")
        return redirect(url_for("login"))

    db = get_db()
    user_id = session["user_id"]

    user = db.execute(
        "SELECT id, nombre, usuario, rol, estado FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not user:
        session.clear()
        flash("Sesión inválida. Inicia sesión nuevamente.", "danger")
        return redirect(url_for("login"))

    if request.method == "POST":
        actual = (request.form.get("password_actual") or "").strip()
        nueva = (request.form.get("password_nueva") or "").strip()
        confirmar = (request.form.get("password_confirmar") or "").strip()

        if not actual or not nueva or not confirmar:
            flash("Completa todos los campos para cambiar contraseña.", "danger")
            return redirect(url_for("perfil"))

        if nueva != confirmar:
            flash("La nueva contraseña y su confirmación no coinciden.", "danger")
            return redirect(url_for("perfil"))

        # Validar seguridad de la nueva contraseña
        errores_password = validar_contrasena_segura(nueva)
        if errores_password:
            flash(" ".join(errores_password), "warning")
            return redirect(url_for("perfil"))

        # Verificar contraseña actual
        user_full = db.execute(
            "SELECT contrasena_hash FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()

        if not user_full or not bcrypt.checkpw(
            actual.encode("utf-8"),
            user_full["contrasena_hash"].encode("utf-8")
        ):
            flash("La contraseña actual no es correcta.", "danger")
            return redirect(url_for("perfil"))

        nuevo_hash = bcrypt.hashpw(
            nueva.encode("utf-8"),
            bcrypt.gensalt()
        ).decode("utf-8")

        db.execute(
            "UPDATE users SET contrasena_hash = ? WHERE id = ?",
            (nuevo_hash, user_id)
        )
        db.commit()

        flash("✅ Contraseña actualizada correctamente.", "success")
        return redirect(url_for("perfil"))

    return render_template("perfil.html", user=user)
    
# =====================================================
# PANEL ADMINISTRADOR
# =====================================================
@app.route("/admin")
def admin_panel():
    if "user_id" not in session:
        return redirect(url_for("login"))

    if session.get("rol") != "admin":
        flash("No tienes permisos.", "danger")
        return redirect(url_for("inicio"))

    db = get_db()

    creditos_pendientes = db.execute("""
        SELECT COUNT(*) AS n
        FROM orders
        WHERE lower(trim(tipo_pago)) IN ('credito', 'crédito')
          AND lower(trim(estado)) = 'pendiente'
    """).fetchone()["n"]

    stock_bajo = db.execute("""
        SELECT COUNT(*) AS n
        FROM products
        WHERE activo = 1
          AND stock_minimo > 0
          AND stock <= stock_minimo
    """).fetchone()["n"]

    usuarios_bloqueados = db.execute("""
        SELECT COUNT(*) AS n
        FROM users
        WHERE estado = 'bloqueado'
    """).fetchone()["n"]

    ordenes_14 = db.execute("""
        SELECT COUNT(*) AS n
        FROM orders
        WHERE date(fecha) >= date('now','-14 day')
          AND date(fecha) <= date('now')
    """).fetchone()["n"]

    return render_template(
        "admin.html",
        kpi_creditos=creditos_pendientes,
        kpi_stock_bajo=stock_bajo,
        kpi_bloqueados=usuarios_bloqueados,
        kpi_ordenes_14=ordenes_14
    )


# =====================================================
# ADMIN: USUARIOS (LISTA)
# =====================================================
@app.route("/admin/users")
def admin_users():
    # Seguridad: solo admin
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    usuarios = db.execute("""
        SELECT id, nombre, usuario, rol, estado
        FROM users
        ORDER BY nombre
    """).fetchall()

    return render_template("admin_users.html", usuarios=usuarios)


# =====================================================
# ADMIN: USUARIOS (AGREGAR)
# =====================================================
@app.route("/admin/users/add", methods=["GET", "POST"])
def admin_user_add():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        usuario = (request.form.get("usuario") or "").strip()
        rol = (request.form.get("rol") or "").strip()
        password = (request.form.get("password") or "").strip()

        if not nombre or not usuario or not password:
            flash("Nombre, usuario y contraseña son obligatorios.", "danger")
            return redirect(url_for("admin_user_add"))

        errores_password = validar_contrasena_segura(password)
        if errores_password:
            flash(" ".join(errores_password), "warning")
            return redirect(url_for("admin_user_add"))

        if rol not in ("admin", "asociado"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("admin_user_add"))

        existe = db.execute(
            "SELECT id FROM users WHERE usuario = ?",
            (usuario,)
        ).fetchone()

        if existe:
            flash("Ese usuario ya existe.", "warning")
            return redirect(url_for("admin_user_add"))

        hash_pw = bcrypt.hashpw(
            password.encode("utf-8"),
            bcrypt.gensalt()
        ).decode("utf-8")

        db.execute(
            """
            INSERT INTO users (nombre, usuario, contrasena_hash, rol, estado)
            VALUES (?, ?, ?, ?, 'activo')
            """,
            (nombre, usuario, hash_pw, rol)
        )
        db.commit()

        flash("Usuario creado correctamente.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_add.html")


# =====================================================
# ADMIN: USUARIOS (EDITAR)
# =====================================================
@app.route("/admin/users/<int:user_id>/edit", methods=["GET", "POST"])
def admin_user_edit(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    u = db.execute(
        "SELECT id, nombre, usuario, rol, estado FROM users WHERE id = ?",
        (user_id,)
    ).fetchone()

    if not u:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        usuario = (request.form.get("usuario") or "").strip()
        rol = (request.form.get("rol") or "").strip()
        estado = (request.form.get("estado") or "").strip()
        new_password = (request.form.get("new_password") or "").strip()

        if not nombre or not usuario:
            flash("Nombre y usuario son obligatorios.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        if rol not in ("admin", "asociado"):
            flash("Rol inválido.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        if estado not in ("activo", "bloqueado"):
            flash("Estado inválido.", "danger")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        existe = db.execute(
            "SELECT id FROM users WHERE usuario = ? AND id <> ?",
            (usuario, user_id)
        ).fetchone()

        if existe:
            flash("Ese usuario ya está en uso por otra cuenta.", "warning")
            return redirect(url_for("admin_user_edit", user_id=user_id))

        db.execute(
            """
            UPDATE users
            SET nombre = ?, usuario = ?, rol = ?, estado = ?
            WHERE id = ?
            """,
            (nombre, usuario, rol, estado, user_id)
        )

        # Si el admin escribe una nueva contraseña, se valida y se actualiza
        if new_password:
            errores_password = validar_contrasena_segura(new_password)

            if errores_password:
                flash(" ".join(errores_password), "warning")
                return redirect(url_for("admin_user_edit", user_id=user_id))

            hash_pw = bcrypt.hashpw(
                new_password.encode("utf-8"),
                bcrypt.gensalt()
            ).decode("utf-8")

            db.execute(
                "UPDATE users SET contrasena_hash = ? WHERE id = ?",
                (hash_pw, user_id)
            )

        db.commit()

        flash("Usuario actualizado correctamente.", "success")
        return redirect(url_for("admin_users"))

    return render_template("admin_user_edit.html", u=u)

# =====================================================
# ADMIN: USUARIOS (BLOQUEAR / DESBLOQUEAR)
# =====================================================
@app.route("/admin/users/<int:user_id>/toggle", methods=["POST"])
def admin_user_toggle(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if session.get("user_id") == user_id:
        flash("No podés bloquear tu propia cuenta.", "warning")
        return redirect(url_for("admin_users"))

    db = get_db()
    u = db.execute("SELECT id, estado FROM users WHERE id = ?", (user_id,)).fetchone()
    if not u:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("admin_users"))

    nuevo_estado = "activo" if u["estado"] == "bloqueado" else "bloqueado"
    db.execute("UPDATE users SET estado = ? WHERE id = ?", (nuevo_estado, user_id))
    db.commit()

    flash(f"Estado actualizado a: {nuevo_estado}", "success")
    return redirect(url_for("admin_users"))


# =====================================================
# MIS COMPRAS (HISTORIAL DEL USUARIO LOGUEADO)
# =====================================================
@app.route("/mis-compras")
def mis_compras():
    if "user_id" not in session:
        flash("Debes iniciar sesión para ver tu historial.", "warning")
        return redirect(url_for("login"))

    user_id = session["user_id"]
    db = get_db()

    pago = (request.args.get("pago") or "").strip().lower()
    pago = pago.replace("é", "e")

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    where = ["o.user_id = ?"]
    params = [user_id]

    if pago in ("contado", "credito"):
        where.append("lower(trim(o.tipo_pago)) = ?")
        params.append(pago)

    if start:
        where.append("date(o.fecha) >= date(?)")
        params.append(start)

    if end:
        where.append("date(o.fecha) <= date(?)")
        params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    orders = db.execute(
        f"""
        SELECT o.id, o.fecha, o.tipo_pago, o.estado, o.total
        FROM orders o
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    total_rango = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "mis_compras.html",
        orders=orders,
        pago=pago,
        start=start,
        end=end,
        total_rango=total_rango
    )


@app.route("/mis-compras/<int:order_id>")
def mis_compras_detalle(order_id):
    if "user_id" not in session:
        flash("Debes iniciar sesión.", "warning")
        return redirect(url_for("login"))

    user_id = session["user_id"]
    db = get_db()

    orden = db.execute(
        """
        SELECT id, fecha, tipo_pago, estado, total
        FROM orders
        WHERE id = ? AND user_id = ?
        """,
        (order_id, user_id)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada o sin permisos.", "danger")
        return redirect(url_for("mis_compras"))

    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        """,
        (order_id,)
    ).fetchall()

    return render_template("mis_compras_detalle.html", orden=orden, items=items)


# =====================================================
# ADMIN: HISTORIAL DE ÓRDENES (lista general con filtros)
# =====================================================
@app.route("/admin/orders")
def admin_orders():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    estado = request.args.get("estado", "").strip()
    pago = request.args.get("pago", "").strip()
    q = request.args.get("q", "").strip()

    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")
    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    if estado:
        where.append("o.estado = ?")
        params.append(estado)

    if pago:
        where.append("o.tipo_pago = ?")
        params.append(pago)

    if q:
        where.append("(u.nombre LIKE ? OR o.nombre_no_asociado LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%"])

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()
    orders = db.execute(
        f"""
        SELECT
            o.*,
            u.nombre AS nombre_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    return render_template(
        "admin_orders.html",
        orders=orders,
        estado=estado,
        pago=pago,
        q=q,
        start=start,
        end=end
    )

# =====================================================
# ADMIN: EXPORTAR HISTORIAL DE ÓRDENES A EXCEL
# Ruta: /admin/orders/export
# - Respeta filtros de admin_orders:
#   estado, pago, q, start, end
# - Si no hay fechas, usa últimos 14 días
# - Genera 2 hojas:
#   1) Órdenes
#   2) Detalle productos
# =====================================================
@app.route("/admin/orders/export")
def admin_orders_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    # ----------------------------
    # Inputs desde URL
    # ----------------------------
    estado = (request.args.get("estado") or "").strip().lower()
    pago = (request.args.get("pago") or "").strip().lower().replace("é", "e")
    q = (request.args.get("q") or "").strip()

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    # ----------------------------
    # Default: últimos 14 días
    # ----------------------------
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    # ----------------------------
    # WHERE dinámico
    # ----------------------------
    where = []
    params = []

    if estado in ("pagada", "pendiente"):
        where.append("lower(trim(o.estado)) = ?")
        params.append(estado)

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    if q:
        where.append("(u.nombre LIKE ? OR o.nombre_no_asociado LIKE ?)")
        params.extend([f"%{q}%", f"%{q}%"])

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    # ----------------------------
    # Resumen
    # ----------------------------
    resumen = db.execute(
        f"""
        SELECT
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        """,
        params
    ).fetchone()

    # ----------------------------
    # Hoja 1: órdenes
    # ----------------------------
    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    # ----------------------------
    # Hoja 2: detalle de productos
    # ----------------------------
    detalle = db.execute(
        f"""
        SELECT
            o.id AS order_id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            o.tipo_pago,
            o.estado,
            o.total AS total_orden,
            p.nombre AS producto,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        JOIN order_items oi ON oi.order_id = o.id
        JOIN products p ON p.id = oi.product_id
        {where_sql}
        ORDER BY o.id DESC, p.nombre ASC
        """,
        params
    ).fetchall()

    # =====================================================
    # CREAR EXCEL
    # =====================================================
    wb = Workbook()

    # =====================================================
    # ESTILOS GENERALES
    # =====================================================
    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)

    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    # =====================================================
    # HOJA 1: ÓRDENES
    # =====================================================
    ws = wb.active
    ws.title = "Órdenes"

    ws.merge_cells("A1:F1")
    ws["A1"] = "Historial de compras - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    # Filtros aplicados
    ws["A3"] = "Desde"
    ws["B3"] = start
    ws["C3"] = "Hasta"
    ws["D3"] = end
    ws["E3"] = "Pago"
    ws["F3"] = pago if pago else "Todos"

    ws["A4"] = "Estado"
    ws["B4"] = estado if estado else "Todos"
    ws["C4"] = "Búsqueda"
    ws["D4"] = q if q else "Sin búsqueda"

    for row in ws.iter_rows(min_row=3, max_row=4, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_center

    # Resumen compacto
    ws["A6"] = "Resumen"
    ws["A6"].font = font_bold
    ws["A6"].fill = fill_title

    ws["A7"] = "Órdenes"
    ws["B7"] = resumen["num_ordenes"]

    ws["A8"] = "Total general"
    ws["B8"] = float(resumen["total_general"])

    ws["A9"] = "Contado"
    ws["B9"] = float(resumen["total_contado"])

    ws["A10"] = "Crédito"
    ws["B10"] = float(resumen["total_credito"])

    for row in ws.iter_rows(min_row=6, max_row=10, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row_num in range(7, 11):
        ws.cell(row=row_num, column=1).font = font_bold

    ws["B8"].number_format = '"₡"#,##0.00'
    ws["B9"].number_format = '"₡"#,##0.00'
    ws["B10"].number_format = '"₡"#,##0.00'

    # Tabla de órdenes
    start_row = 12
    headers = ["Orden", "Fecha", "Comprador", "Pago", "Estado", "Total"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for o in orders:
        ws.cell(row=current_row, column=1).value = f"#{o['id']}"
        ws.cell(row=current_row, column=2).value = o["fecha"]
        ws.cell(row=current_row, column=3).value = o["comprador"]
        ws.cell(row=current_row, column=4).value = o["tipo_pago"]
        ws.cell(row=current_row, column=5).value = o["estado"]
        ws.cell(row=current_row, column=6).value = float(o["total"])

        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=6).number_format = '"₡"#,##0.00'

        current_row += 1

    if not orders:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1).value = "Sin datos para los filtros seleccionados."
        ws.cell(row=current_row, column=1).alignment = alignment_center

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 16

    ws.freeze_panes = "A13"
    ws.auto_filter.ref = f"A{start_row}:F{max(start_row, current_row - 1)}"

    # =====================================================
    # HOJA 2: DETALLE PRODUCTOS
    # =====================================================
    ws2 = wb.create_sheet("Detalle productos")

    ws2.merge_cells("A1:J1")
    ws2["A1"] = "Detalle de productos por orden - ASERVE"
    ws2["A1"].font = font_title
    ws2["A1"].fill = fill_title
    ws2["A1"].alignment = alignment_left

    headers_detalle = [
        "Orden",
        "Fecha",
        "Comprador",
        "Pago",
        "Estado",
        "Total orden",
        "Producto",
        "Cantidad",
        "Precio unitario",
        "Subtotal"
    ]

    start_row_detalle = 3

    for col_num, header in enumerate(headers_detalle, start=1):
        cell = ws2.cell(row=start_row_detalle, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row_detalle + 1

    for d in detalle:
        ws2.cell(row=current_row, column=1).value = f"#{d['order_id']}"
        ws2.cell(row=current_row, column=2).value = d["fecha"]
        ws2.cell(row=current_row, column=3).value = d["comprador"]
        ws2.cell(row=current_row, column=4).value = d["tipo_pago"]
        ws2.cell(row=current_row, column=5).value = d["estado"]
        ws2.cell(row=current_row, column=6).value = float(d["total_orden"])
        ws2.cell(row=current_row, column=7).value = d["producto"]
        ws2.cell(row=current_row, column=8).value = int(d["cantidad"])
        ws2.cell(row=current_row, column=9).value = float(d["precio_unitario"])
        ws2.cell(row=current_row, column=10).value = float(d["subtotal"])

        for col in range(1, 11):
            cell = ws2.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws2.cell(row=current_row, column=6).number_format = '"₡"#,##0.00'
        ws2.cell(row=current_row, column=9).number_format = '"₡"#,##0.00'
        ws2.cell(row=current_row, column=10).number_format = '"₡"#,##0.00'

        current_row += 1

    if not detalle:
        ws2.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        ws2.cell(row=current_row, column=1).value = "Sin detalle de productos para los filtros seleccionados."
        ws2.cell(row=current_row, column=1).alignment = alignment_center

    column_widths_detalle = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 15,
        "E": 15,
        "F": 16,
        "G": 35,
        "H": 12,
        "I": 18,
        "J": 16,
    }

    for col_letter, width in column_widths_detalle.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A4"
    ws2.auto_filter.ref = f"A{start_row_detalle}:J{max(start_row_detalle, current_row - 1)}"

    # =====================================================
    # DESCARGA
    # =====================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"historial_compras_{start}_a_{end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: DETALLE DE ORDEN
# =====================================================
@app.route("/admin/orders/<int:order_id>")
def admin_order_detail(order_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    orden = db.execute(
        """
        SELECT
            o.*,
            u.nombre AS nombre_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.id = ?
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada.", "danger")
        return redirect(url_for("admin_orders"))

    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        """,
        (order_id,)
    ).fetchall()

    from_page = request.args.get("from_page", "orders")
    back_user_id = request.args.get("user_id", "").strip()

    return render_template(
        "admin_order_detail.html",
        orden=orden,
        items=items,
        from_page=from_page,
        back_user_id=back_user_id
    )


# =====================================================
# ADMIN: LISTA DE COMPRADORES
# =====================================================
@app.route("/admin/buyers")
def admin_buyers():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    registrados = db.execute(
        """
        SELECT
            ('u:' || u.id) AS buyer_key,
            u.nombre AS comprador,
            'registrado' AS tipo,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_monto
        FROM orders o
        JOIN users u ON u.id = o.user_id
        GROUP BY u.id, u.nombre
        """
    ).fetchall()

    no_asociados = db.execute(
        """
        SELECT
            ('na:' || o.nombre_no_asociado) AS buyer_key,
            o.nombre_no_asociado AS comprador,
            'no_asociado' AS tipo,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_monto
        FROM orders o
        WHERE o.user_id IS NULL
          AND o.nombre_no_asociado IS NOT NULL
          AND TRIM(o.nombre_no_asociado) <> ''
        GROUP BY o.nombre_no_asociado
        """
    ).fetchall()

    buyers = list(registrados) + list(no_asociados)
    buyers.sort(key=lambda x: (-float(x["total_monto"]), x["comprador"]))

    return render_template("admin_buyers.html", buyers=buyers)


# =====================================================
# ADMIN: HISTORIAL POR COMPRADOR
# - Muestra órdenes de un comprador
# - Filtra automáticamente últimos 14 días si no se envían fechas
# =====================================================
@app.route("/admin/buyer")
def admin_buyer_history():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    key = request.args.get("key", "").strip()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()

    if not key:
        flash("Comprador inválido.", "danger")
        return redirect(url_for("admin_buyers"))

    # =====================================================
    # DEFAULT: últimos 14 días
    # Si no se envían fechas, se carga automáticamente
    # desde hace 13 días hasta hoy.
    # Ejemplo: si hoy es 26, muestra del 13 al 26.
    # =====================================================
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    db = get_db()

    where = []
    params = []

    comprador_nombre = ""
    comprador_tipo = ""

    # =====================================================
    # Comprador registrado
    # =====================================================
    if key.startswith("u:"):
        comprador_tipo = "registrado"
        user_id = key.split(":", 1)[1]

        where.append("o.user_id = ?")
        params.append(user_id)

        u = db.execute(
            "SELECT nombre FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()

        comprador_nombre = u["nombre"] if u else "Usuario"

    # =====================================================
    # Comprador no asociado
    # =====================================================
    elif key.startswith("na:"):
        comprador_tipo = "no_asociado"
        nombre = key.split(":", 1)[1]

        where.append("o.user_id IS NULL")
        where.append("o.nombre_no_asociado = ?")
        params.append(nombre)

        comprador_nombre = nombre

    else:
        flash("Comprador inválido.", "danger")
        return redirect(url_for("admin_buyers"))

    # =====================================================
    # Filtro de fechas
    # Siempre se aplica, porque si no vienen fechas,
    # arriba ya se asignaron los últimos 14 días.
    # =====================================================
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    total_rango = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "admin_buyer_history.html",
        key=key,
        comprador_nombre=comprador_nombre,
        comprador_tipo=comprador_tipo,
        orders=orders,
        total_rango=total_rango,
        start=start,
        end=end
    )

# =====================================================
# ADMIN: CRÉDITOS (RESUMEN)
# =====================================================
@app.route("/admin/credits")
def admin_credits():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    creditos = db.execute(
        """
        SELECT
            u.id AS user_id,
            u.nombre AS comprador,
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS deuda_total
        FROM orders o
        JOIN users u ON u.id = o.user_id
        WHERE
            replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito'
            AND lower(trim(o.estado)) = 'pendiente'
        GROUP BY u.id, u.nombre
        ORDER BY deuda_total DESC, u.nombre ASC
        """
    ).fetchall()

    return render_template("admin_credits.html", creditos=creditos)


# =====================================================
# ADMIN: CRÉDITOS POR COLABORADOR
# =====================================================
@app.route("/admin/credits/user/<int:user_id>")
def admin_credits_user(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    user = db.execute("SELECT id, nombre FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        flash("Colaborador no encontrado.", "danger")
        return redirect(url_for("admin_credits"))

    orders = db.execute(
        """
        SELECT id, fecha, total
        FROM orders
        WHERE user_id = ?
          AND replace(lower(trim(tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(estado)) = 'pendiente'
        ORDER BY fecha ASC
        """,
        (user_id,)
    ).fetchall()

    deuda_total = sum(float(o["total"]) for o in orders) if orders else 0

    return render_template(
        "admin_credits_user.html",
        user=user,
        orders=orders,
        deuda_total=deuda_total
    )


# =====================================================
# ADMIN: PAGAR CRÉDITOS EN BLOQUE
# =====================================================
@app.route("/admin/credits/user/<int:user_id>/pay-bulk", methods=["POST"])
def admin_credits_pay_bulk(user_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    order_ids = request.form.getlist("order_ids")

    if not order_ids:
        flash("No seleccionaste ninguna orden.", "warning")
        return redirect(url_for("admin_credits_user", user_id=user_id))

    try:
        order_ids_int = [int(x) for x in order_ids]
    except ValueError:
        flash("Selección inválida.", "danger")
        return redirect(url_for("admin_credits_user", user_id=user_id))

    db = get_db()
    placeholders = ",".join(["?"] * len(order_ids_int))

    cur = db.execute(
        f"""
        UPDATE orders
        SET estado = 'pagada'
        WHERE user_id = ?
          AND lower(trim(tipo_pago)) IN ('credito', 'crédito')
          AND lower(trim(estado)) = 'pendiente'
          AND id IN ({placeholders})
        """,
        [user_id] + order_ids_int
    )

    db.commit()

    flash(f"Se marcaron como pagadas {cur.rowcount} orden(es).", "success")
    return redirect(url_for("admin_credits_user", user_id=user_id))


# =====================================================
# ADMIN: PAGAR CRÉDITO INDIVIDUAL
# =====================================================
@app.route("/admin/credits/pay/<int:order_id>", methods=["POST"])
def admin_credits_pay(order_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    orden = db.execute(
        """
        SELECT id, user_id
        FROM orders
        WHERE id = ?
          AND replace(lower(trim(tipo_pago)), 'é', 'e') = 'credito'
          AND lower(trim(estado)) = 'pendiente'
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Esa orden no existe o ya fue pagada.", "warning")
        return redirect(url_for("admin_credits"))

    db.execute("UPDATE orders SET estado = 'pagada' WHERE id = ?", (order_id,))
    db.commit()

    flash(f"Orden #{order_id} marcada como pagada.", "success")
    return redirect(url_for("admin_credits_user", user_id=orden["user_id"]))


# =====================================================
# ADMIN: STOCK
# =====================================================
@app.route("/admin/stock")
def admin_stock():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    only_low = (request.args.get("only_low") or "").strip() == "1"

    db = get_db()
    productos = db.execute(
        """
        SELECT id, nombre, stock, stock_minimo, activo
        FROM products
        ORDER BY nombre ASC
        """
    ).fetchall()

    bajos = []
    for p in productos:
        if int(p["activo"]) == 1 and int(p["stock_minimo"]) > 0 and int(p["stock"]) <= int(p["stock_minimo"]):
            bajos.append(p)

    if only_low:
        productos = bajos

    return render_template(
        "admin_stock.html",
        productos=productos,
        bajos_count=len(bajos),
        only_low=only_low
    )


# =====================================================
# ADMIN: PRODUCTOS
# =====================================================
@app.route("/admin/products")
def admin_products():
    # Seguridad: solo admin
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()

    productos = db.execute("""
        SELECT id, nombre, precio, stock, stock_minimo, activo, image_filename
        FROM products
        ORDER BY nombre ASC
    """).fetchall()

    bajos = [
        p for p in productos
        if int(p["activo"]) == 1
        and int(p["stock_minimo"] or 0) > 0
        and int(p["stock"] or 0) <= int(p["stock_minimo"] or 0)
    ]

    return render_template("admin_products.html", productos=productos, bajos=bajos)


@app.route("/admin/products/add", methods=["GET", "POST"])
def admin_product_add():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        precio = (request.form.get("precio") or "").strip()
        stock = (request.form.get("stock") or "").strip()
        stock_minimo = (request.form.get("stock_minimo") or "0").strip()
        activo = 1 if request.form.get("activo") == "on" else 0

        imagen = request.files.get("imagen")

        if not nombre or not precio or not stock:
            flash("Nombre, precio y stock son obligatorios.", "danger")
            return redirect(url_for("admin_product_add"))

        try:
            precio_num = float(precio)
            stock_num = int(stock)
            stock_minimo_num = int(stock_minimo)
            if precio_num < 0 or stock_num < 0 or stock_minimo_num < 0:
                flash("Precio/stock no pueden ser negativos.", "danger")
                return redirect(url_for("admin_product_add"))
        except ValueError:
            flash("Precio debe ser número y stock/stock mínimo deben ser enteros.", "danger")
            return redirect(url_for("admin_product_add"))

        image_filename = None
        if imagen and imagen.filename:
            if not allowed_file(imagen.filename):
                flash("Formato de imagen no permitido.", "danger")
                return redirect(url_for("admin_product_add"))

            filename = secure_filename(imagen.filename)
            unique_name = f"{int(datetime.now().timestamp())}_{filename}"
            os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
            imagen.save(os.path.join(app.config["UPLOAD_FOLDER"], unique_name))
            image_filename = unique_name

        db = get_db()
        db.execute("""
            INSERT INTO products (nombre, precio, stock, stock_minimo, activo, image_filename)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (nombre, precio_num, stock_num, stock_minimo_num, activo, image_filename))
        db.commit()

        flash("✅ Producto agregado.", "success")
        return redirect(url_for("admin_products"))

    return render_template("admin_product_add.html")


@app.route("/admin/products/<int:product_id>/edit", methods=["GET", "POST"])
def admin_product_edit(product_id):
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    db = get_db()
    p = db.execute("""
        SELECT id, nombre, precio, stock, stock_minimo, activo, image_filename
        FROM products
        WHERE id = ?
    """, (product_id,)).fetchone()

    if not p:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("admin_products"))

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        precio = (request.form.get("precio") or "").strip()
        stock = (request.form.get("stock") or "").strip()
        stock_minimo = (request.form.get("stock_minimo") or "0").strip()
        activo = 1 if request.form.get("activo") == "on" else 0

        imagen = request.files.get("imagen")

        if not nombre or not precio or not stock:
            flash("Nombre, precio y stock son obligatorios.", "danger")
            return redirect(url_for("admin_product_edit", product_id=product_id))

        try:
            precio_num = float(precio)
            stock_num = int(stock)
            stock_minimo_num = int(stock_minimo)
            if precio_num < 0 or stock_num < 0 or stock_minimo_num < 0:
                flash("Precio/stock no pueden ser negativos.", "danger")
                return redirect(url_for("admin_product_edit", product_id=product_id))
        except ValueError:
            flash("Precio debe ser número y stock/stock mínimo deben ser enteros.", "danger")
            return redirect(url_for("admin_product_edit", product_id=product_id))

        image_filename = p["image_filename"]
        if imagen and imagen.filename:
            if not allowed_file(imagen.filename):
                flash("Formato de imagen no permitido.", "danger")
                return redirect(url_for("admin_product_edit", product_id=product_id))

            filename = secure_filename(imagen.filename)
            unique_name = f"{int(datetime.now().timestamp())}_{filename}"
            os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
            imagen.save(os.path.join(app.config["UPLOAD_FOLDER"], unique_name))
            image_filename = unique_name

        db.execute("""
            UPDATE products
            SET nombre = ?, precio = ?, stock = ?, stock_minimo = ?, activo = ?, image_filename = ?
            WHERE id = ?
        """, (nombre, precio_num, stock_num, stock_minimo_num, activo, image_filename, product_id))
        db.commit()

        flash("✅ Producto actualizado.", "success")
        return redirect(url_for("admin_products"))

    return render_template("admin_product_edit.html", p=p)


# =====================================================
# ADMIN: REPORTE VENTAS POR PERIODO
# =====================================================
@app.route("/admin/reports/sales")
def admin_report_sales():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    pago = (request.args.get("pago") or "").strip().lower()
    pago = pago.replace("é", "e")

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    resumen = db.execute(
        f"""
        SELECT
            COUNT(*) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    orders = db.execute(
        f"""
        SELECT
            o.id, o.fecha, o.tipo_pago, o.estado, o.total,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    return render_template(
        "admin_report_sales.html",
        start=start,
        end=end,
        pago=pago,
        resumen=resumen,
        orders=orders
    )


# =====================================================
# ADMIN: EXPORTAR REPORTE DE VENTAS A EXCEL
# =====================================================
@app.route("/admin/reports/sales/export")
def admin_report_sales_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    pago = (request.args.get("pago") or "").strip().lower()
    pago = pago.replace("é", "e")

    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    where = []
    params = []

    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    if pago in ("contado", "credito"):
        where.append("replace(lower(trim(o.tipo_pago)), 'é', 'e') = ?")
        params.append(pago)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    resumen = db.execute(
        f"""
        SELECT
            COUNT(*) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_general,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'contado' THEN o.total
                ELSE 0
            END), 0) AS total_contado,
            COALESCE(SUM(CASE
                WHEN replace(lower(trim(o.tipo_pago)), 'é', 'e') = 'credito' THEN o.total
                ELSE 0
            END), 0) AS total_credito
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    orders = db.execute(
        f"""
        SELECT
            o.id,
            o.fecha,
            COALESCE(u.nombre, o.nombre_no_asociado, 'Sin nombre') AS comprador,
            o.tipo_pago,
            o.estado,
            o.total
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        {where_sql}
        ORDER BY o.id DESC
        """,
        params
    ).fetchall()

    # Crear archivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte ventas"

    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

        # Encabezado del reporte
    ws.merge_cells("A1:F1")
    ws["A1"] = "Reporte de ventas - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    # Filtros aplicados
    ws["A3"] = "Desde"
    ws["B3"] = start
    ws["C3"] = "Hasta"
    ws["D3"] = end
    ws["E3"] = "Tipo de pago"
    ws["F3"] = pago if pago else "Todos"

    for cell in ws[3]:
        cell.border = thin_border
        cell.alignment = alignment_center
        cell.font = font_bold

    # Resumen compacto
    ws["A5"] = "Resumen"
    ws["A5"].font = font_bold
    ws["A5"].fill = fill_title
    ws["A5"].alignment = alignment_left

    ws["A6"] = "Órdenes"
    ws["B6"] = resumen["num_ordenes"]

    ws["A7"] = "Total general"
    ws["B7"] = float(resumen["total_general"])

    ws["A8"] = "Contado"
    ws["B8"] = float(resumen["total_contado"])

    ws["A9"] = "Crédito"
    ws["B9"] = float(resumen["total_credito"])

    for row in ws.iter_rows(min_row=5, max_row=9, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row in range(6, 10):
        ws.cell(row=row, column=1).font = font_bold

    ws["B7"].number_format = '"₡"#,##0.00'
    ws["B8"].number_format = '"₡"#,##0.00'
    ws["B9"].number_format = '"₡"#,##0.00'
    for row in ws.iter_rows(min_row=5, max_row=6, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_center
            if cell.column in (1, 3, 5):
                cell.font = font_bold

    ws["D5"].number_format = '"₡"#,##0.00'
    ws["F5"].number_format = '"₡"#,##0.00'
    ws["B6"].number_format = '##0'

    # Tabla de órdenes
    start_row = 11
    headers = ["Orden", "Fecha", "Comprador", "Pago", "Estado", "Total"]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for o in orders:
        ws.cell(row=current_row, column=1).value = f"#{o['id']}"
        ws.cell(row=current_row, column=2).value = o["fecha"]
        ws.cell(row=current_row, column=3).value = o["comprador"]
        ws.cell(row=current_row, column=4).value = o["tipo_pago"]
        ws.cell(row=current_row, column=5).value = o["estado"]
        ws.cell(row=current_row, column=6).value = float(o["total"])

        for col in range(1, 7):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=6).number_format = '"₡"#,##0.00'

        current_row += 1

    if not orders:
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=1).value = "Sin datos para el filtro seleccionado."
        ws.cell(row=current_row, column=1).alignment = alignment_center

    column_widths = {
        "A": 12,
        "B": 22,
        "C": 35,
        "D": 15,
        "E": 15,
        "F": 16,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A12"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"reporte_ventas_{start}_a_{end}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: EXPORTAR HISTORIAL POR COMPRADOR A EXCEL
# Ruta: /admin/buyer/export?key=...&start=...&end=...
# - Exporta todas las órdenes de un comprador
# - Incluye detalle de artículos comprados por orden
# =====================================================
@app.route("/admin/buyer/export")
def admin_buyer_history_export():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    key = request.args.get("key", "").strip()
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()

    if not key:
        flash("Comprador inválido para exportar.", "danger")
        return redirect(url_for("admin_buyers"))

    db = get_db()

    where = []
    params = []

    comprador_nombre = ""
    comprador_tipo = ""

    # ----------------------------
    # Identificar comprador
    # ----------------------------
    if key.startswith("u:"):
        comprador_tipo = "registrado"
        user_id = key.split(":", 1)[1]

        where.append("o.user_id = ?")
        params.append(user_id)

        u = db.execute(
            "SELECT nombre FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()

        comprador_nombre = u["nombre"] if u else "Usuario"

    elif key.startswith("na:"):
        comprador_tipo = "no asociado"
        nombre = key.split(":", 1)[1]

        where.append("o.user_id IS NULL")
        where.append("o.nombre_no_asociado = ?")
        params.append(nombre)

        comprador_nombre = nombre

    else:
        flash("Comprador inválido para exportar.", "danger")
        return redirect(url_for("admin_buyers"))

    # ----------------------------
    # Filtro por fechas
    # ----------------------------
    if start:
        where.append("date(o.fecha) >= date(?)")
        params.append(start)

    if end:
        where.append("date(o.fecha) <= date(?)")
        params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    # ----------------------------
    # Resumen del comprador
    # ----------------------------
    resumen = db.execute(
        f"""
        SELECT
            COUNT(o.id) AS num_ordenes,
            COALESCE(SUM(o.total), 0) AS total_acumulado
        FROM orders o
        {where_sql}
        """,
        params
    ).fetchone()

    # ----------------------------
    # Detalle con productos comprados
    # ----------------------------
    rows = db.execute(
        f"""
        SELECT
            o.id AS order_id,
            o.fecha,
            o.tipo_pago,
            o.estado,
            o.total AS total_orden,
            COALESCE(p.nombre, 'Sin producto') AS producto,
            COALESCE(oi.cantidad, 0) AS cantidad,
            COALESCE(oi.precio_unitario, 0) AS precio_unitario,
            COALESCE((oi.cantidad * oi.precio_unitario), 0) AS subtotal
        FROM orders o
        LEFT JOIN order_items oi ON oi.order_id = o.id
        LEFT JOIN products p ON p.id = oi.product_id
        {where_sql}
        ORDER BY o.fecha DESC, o.id DESC, p.nombre ASC
        """,
        params
    ).fetchall()

    # =====================================================
    # CREAR EXCEL
    # =====================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Historial comprador"

    # Estilos
    fill_header = PatternFill("solid", fgColor="1F2328")
    fill_title = PatternFill("solid", fgColor="EAF2F8")
    font_header = Font(color="FFFFFF", bold=True)
    font_title = Font(bold=True, size=14)
    font_bold = Font(bold=True)
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9DEE3"),
        right=Side(style="thin", color="D9DEE3"),
        top=Side(style="thin", color="D9DEE3"),
        bottom=Side(style="thin", color="D9DEE3")
    )

    # ----------------------------
    # Título
    # ----------------------------
    ws.merge_cells("A1:I1")
    ws["A1"] = "Historial de compras por comprador - ASERVE"
    ws["A1"].font = font_title
    ws["A1"].fill = fill_title
    ws["A1"].alignment = alignment_left

    # ----------------------------
    # Datos del comprador
    # ----------------------------
    ws["A3"] = "Comprador"
    ws["B3"] = comprador_nombre

    ws["A4"] = "Tipo"
    ws["B4"] = comprador_tipo

    ws["A5"] = "Desde"
    ws["B5"] = start if start else "Inicio"

    ws["A6"] = "Hasta"
    ws["B6"] = end if end else "Final"

    ws["A8"] = "Resumen"
    ws["A8"].font = font_bold
    ws["A8"].fill = fill_title

    ws["A9"] = "Órdenes"
    ws["B9"] = resumen["num_ordenes"]

    ws["A10"] = "Total acumulado"
    ws["B10"] = float(resumen["total_acumulado"])
    ws["B10"].number_format = '"₡"#,##0.00'

    for row in ws.iter_rows(min_row=3, max_row=10, min_col=1, max_col=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = alignment_left

    for row in range(3, 11):
        ws.cell(row=row, column=1).font = font_bold

    # ----------------------------
    # Tabla de detalle
    # ----------------------------
    start_row = 12

    headers = [
        "Orden",
        "Fecha",
        "Tipo de pago",
        "Estado",
        "Total orden",
        "Producto",
        "Cantidad",
        "Precio unitario",
        "Subtotal"
    ]

    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header
        cell.fill = fill_header
        cell.font = font_header
        cell.alignment = alignment_center
        cell.border = thin_border

    current_row = start_row + 1

    for r in rows:
        ws.cell(row=current_row, column=1).value = f"#{r['order_id']}"
        ws.cell(row=current_row, column=2).value = r["fecha"]
        ws.cell(row=current_row, column=3).value = r["tipo_pago"]
        ws.cell(row=current_row, column=4).value = r["estado"]
        ws.cell(row=current_row, column=5).value = float(r["total_orden"])
        ws.cell(row=current_row, column=6).value = r["producto"]
        ws.cell(row=current_row, column=7).value = int(r["cantidad"])
        ws.cell(row=current_row, column=8).value = float(r["precio_unitario"])
        ws.cell(row=current_row, column=9).value = float(r["subtotal"])

        for col in range(1, 10):
            cell = ws.cell(row=current_row, column=col)
            cell.border = thin_border
            cell.alignment = alignment_left

        ws.cell(row=current_row, column=5).number_format = '"₡"#,##0.00'
        ws.cell(row=current_row, column=8).number_format = '"₡"#,##0.00'
        ws.cell(row=current_row, column=9).number_format = '"₡"#,##0.00'

        current_row += 1

    if not rows:
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=9
        )
        ws.cell(row=current_row, column=1).value = "Sin datos para este comprador."
        ws.cell(row=current_row, column=1).alignment = alignment_center
        ws.cell(row=current_row, column=1).border = thin_border

    # ----------------------------
    # Ajustes visuales
    # ----------------------------
    column_widths = {
        "A": 12,
        "B": 22,
        "C": 16,
        "D": 14,
        "E": 16,
        "F": 35,
        "G": 12,
        "H": 18,
        "I": 16,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.freeze_panes = "A13"
    ws.auto_filter.ref = f"A{start_row}:I{current_row - 1}"

    # ----------------------------
    # Descargar archivo
    # ----------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = secure_filename(comprador_nombre) or "comprador"
    filename = f"historial_comprador_{safe_name}.xlsx"

    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# =====================================================
# ADMIN: TOP PRODUCTOS
# Ruta: /admin/reports/top-products
# - Muestra los productos más vendidos
# - Filtra automáticamente últimos 14 días si no se envían fechas
# =====================================================
@app.route("/admin/reports/top-products")
def admin_report_top_products():
    if "user_id" not in session or session.get("rol") != "admin":
        return redirect(url_for("login"))

    # ----------------------------
    # Fechas desde formulario
    # ----------------------------
    start = (request.args.get("start") or "").strip()
    end = (request.args.get("end") or "").strip()

    # ----------------------------
    # Default: últimos 14 días
    # Ejemplo: si hoy es 21, filtra del 8 al 21
    # ----------------------------
    hoy = datetime.now().date()
    hace_14 = hoy - timedelta(days=13)

    if not start:
        start = hace_14.strftime("%Y-%m-%d")

    if not end:
        end = hoy.strftime("%Y-%m-%d")

    # ----------------------------
    # WHERE dinámico
    # ----------------------------
    where = []
    params = []

    # Siempre se aplica el rango de fechas
    where.append("date(o.fecha) >= date(?)")
    params.append(start)

    where.append("date(o.fecha) <= date(?)")
    params.append(end)

    where_sql = "WHERE " + " AND ".join(where)

    db = get_db()

    # ----------------------------
    # Consulta top productos
    # ----------------------------
    top = db.execute(
        f"""
        SELECT
            p.id,
            p.nombre,
            SUM(oi.cantidad) AS unidades,
            COALESCE(SUM(oi.cantidad * oi.precio_unitario), 0) AS monto
        FROM order_items oi
        JOIN orders o ON o.id = oi.order_id
        JOIN products p ON p.id = oi.product_id
        {where_sql}
        GROUP BY p.id, p.nombre
        ORDER BY unidades DESC, monto DESC
        LIMIT 20
        """,
        params
    ).fetchall()

    return render_template(
        "admin_report_top_products.html",
        start=start,
        end=end,
        top=top
    )

# =====================================================
# CATÁLOGO
# =====================================================
@app.route("/catalogo")
def catalogo():
    db = get_db()
    q = (request.args.get("q") or "").strip()

    if q:
        productos = db.execute(
            """
            SELECT * FROM products
            WHERE activo = 1
              AND stock > 0
              AND nombre LIKE ?
            ORDER BY nombre ASC
            """,
            (f"%{q}%",)
        ).fetchall()
    else:
        productos = db.execute(
            """
            SELECT * FROM products
            WHERE activo = 1 AND stock > 0
            ORDER BY nombre ASC
            """
        ).fetchall()

    return render_template("catalogo.html", productos=productos, q=q)


# =====================================================
# HELPER CARRITO
# =====================================================
def get_cart():
    if "cart" not in session:
        session["cart"] = {}
    return session["cart"]


# =====================================================
# CARRITO
# =====================================================
@app.route("/carrito")
def ver_carrito():
    cart = get_cart()

    if not cart:
        return render_template("carrito.html", items=[], total=0)

    db = get_db()
    ids = list(cart.keys())
    placeholders = ",".join(["?"] * len(ids))

    productos = db.execute(
        f"SELECT * FROM products WHERE id IN ({placeholders})",
        ids
    ).fetchall()

    items = []
    total = 0

    for p in productos:
        pid = str(p["id"])
        cantidad = int(cart.get(pid, 0))
        subtotal = float(p["precio"]) * cantidad
        total += subtotal

        items.append({
            "id": p["id"],
            "nombre": p["nombre"],
            "precio": float(p["precio"]),
            "stock": int(p["stock"]),
            "image_filename": p["image_filename"],
            "cantidad": cantidad,
            "subtotal": subtotal
        })

    return render_template("carrito.html", items=items, total=total)


@app.route("/carrito/add/<int:product_id>", methods=["POST"])
def carrito_add(product_id):
    db = get_db()

    p = db.execute(
        "SELECT * FROM products WHERE id = ? AND activo = 1",
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no disponible.", "danger")
        return redirect(url_for("catalogo"))

    cart = get_cart()
    pid = str(product_id)
    cantidad_actual = int(cart.get(pid, 0))

    if cantidad_actual + 1 > int(p["stock"]):
        flash("No hay suficiente stock para agregar más.", "warning")
        return redirect(url_for("catalogo"))

    cart[pid] = cantidad_actual + 1
    session["cart"] = cart
    flash("Producto agregado al carrito.", "success")
    return redirect(url_for("catalogo"))


@app.route("/carrito/inc/<int:product_id>", methods=["POST"])
def carrito_inc(product_id):
    db = get_db()
    p = db.execute(
        "SELECT * FROM products WHERE id = ? AND activo = 1",
        (product_id,)
    ).fetchone()

    if not p:
        flash("Producto no disponible.", "danger")
        return redirect(url_for("ver_carrito"))

    cart = get_cart()
    pid = str(product_id)
    cantidad_actual = int(cart.get(pid, 0))

    if cantidad_actual + 1 > int(p["stock"]):
        flash("No hay suficiente stock.", "warning")
        return redirect(url_for("ver_carrito"))

    cart[pid] = cantidad_actual + 1
    session["cart"] = cart
    return redirect(url_for("ver_carrito"))


@app.route("/carrito/dec/<int:product_id>", methods=["POST"])
def carrito_dec(product_id):
    cart = get_cart()
    pid = str(product_id)
    cantidad_actual = int(cart.get(pid, 0))

    if cantidad_actual <= 1:
        cart.pop(pid, None)
    else:
        cart[pid] = cantidad_actual - 1

    session["cart"] = cart
    return redirect(url_for("ver_carrito"))


@app.route("/carrito/remove/<int:product_id>", methods=["POST"])
def carrito_remove(product_id):
    cart = get_cart()
    cart.pop(str(product_id), None)
    session["cart"] = cart
    return redirect(url_for("ver_carrito"))


@app.route("/carrito/clear", methods=["POST"])
def carrito_clear():
    session["cart"] = {}
    flash("Carrito vaciado.", "info")
    return redirect(url_for("ver_carrito"))


# =====================================================
# CHECKOUT
# =====================================================
@app.route("/checkout", methods=["GET", "POST"])
def checkout():
    cart = get_cart()
    if not cart:
        flash("Tu carrito está vacío.", "info")
        return redirect(url_for("catalogo"))

    db = get_db()

    ids = list(cart.keys())
    placeholders = ",".join(["?"] * len(ids))

    productos = db.execute(
        f"SELECT * FROM products WHERE id IN ({placeholders})",
        ids
    ).fetchall()

    items = []
    total = 0

    for p in productos:
        pid = str(p["id"])
        cantidad = int(cart.get(pid, 0))

        if int(p["activo"]) != 1:
            flash(f"El producto '{p['nombre']}' no está disponible.", "danger")
            return redirect(url_for("ver_carrito"))

        if cantidad > int(p["stock"]):
            flash(f"No hay stock suficiente de '{p['nombre']}'.", "warning")
            return redirect(url_for("ver_carrito"))

        subtotal = float(p["precio"]) * cantidad
        total += subtotal

        items.append({
            "id": p["id"],
            "nombre": p["nombre"],
            "precio": float(p["precio"]),
            "cantidad": cantidad,
            "subtotal": subtotal
        })

    user_id = session.get("user_id")
    rol = session.get("rol")
    tipo_usuario = "asociado" if user_id else "no_asociado"

    if request.method == "GET":
        es_no_asociado = True if not user_id else False
        can_credit = True if (user_id and rol in ("admin", "asociado")) else False

        return render_template(
            "checkout.html",
            items=items,
            total=total,
            rol=rol,
            es_no_asociado=es_no_asociado,
            can_credit=can_credit
        )

    tipo_pago = norm_text(request.form.get("tipo_pago"))
    if tipo_pago not in ("contado", "credito"):
        flash("Tipo de pago inválido.", "danger")
        return redirect(url_for("checkout"))

    nombre_no_asociado = (request.form.get("nombre_no_asociado") or "").strip()

    if tipo_usuario == "no_asociado":
        if tipo_pago != "contado":
            flash("Solo usuarios registrados pueden comprar a crédito.", "danger")
            return redirect(url_for("checkout"))
        if not nombre_no_asociado:
            flash("Debes escribir tu nombre para finalizar la compra.", "danger")
            return redirect(url_for("checkout"))

    if tipo_usuario == "asociado" and tipo_pago == "credito":
        if rol not in ("admin", "asociado"):
            flash("Tu usuario no tiene permiso para comprar a crédito.", "danger")
            return redirect(url_for("checkout"))

    estado = "pagada" if tipo_pago == "contado" else "pendiente"
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cur = db.execute(
        """
        INSERT INTO orders (fecha, tipo_usuario, user_id, nombre_no_asociado, tipo_pago, total, estado)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (
            fecha,
            tipo_usuario,
            user_id if tipo_usuario == "asociado" else None,
            nombre_no_asociado if tipo_usuario == "no_asociado" else None,
            tipo_pago,
            float(total),
            estado
        )
    )
    order_id = cur.lastrowid

    for it in items:
        db.execute(
            """
            INSERT INTO order_items (order_id, product_id, cantidad, precio_unitario)
            VALUES (?, ?, ?, ?)
            """,
            (order_id, it["id"], it["cantidad"], float(it["precio"]))
        )

        db.execute(
            "UPDATE products SET stock = stock - ? WHERE id = ?",
            (it["cantidad"], it["id"])
        )

        db.execute(
            """
            INSERT INTO stock_movements (product_id, cambio_stock, motivo, fecha, order_id)
            VALUES (?, ?, ?, ?, ?)
            """,
            (it["id"], -it["cantidad"], "Venta", fecha, order_id)
        )

    db.commit()

    session["cart"] = {}
    flash(f"✅ Compra realizada. Orden #{order_id} creada.", "success")
    return redirect(url_for("order_success", order_id=order_id))


# =====================================================
# ORDEN SUCCESS
# =====================================================
@app.route("/orden/<int:order_id>")
def order_success(order_id):
    db = get_db()

    orden = db.execute(
        """
        SELECT
            o.*,
            u.nombre AS nombre_usuario
        FROM orders o
        LEFT JOIN users u ON u.id = o.user_id
        WHERE o.id = ?
        """,
        (order_id,)
    ).fetchone()

    if not orden:
        flash("Orden no encontrada.", "danger")
        return redirect(url_for("catalogo"))

    comprador = orden["nombre_usuario"] if orden["user_id"] else orden["nombre_no_asociado"]

    items = db.execute(
        """
        SELECT
            p.nombre,
            oi.cantidad,
            oi.precio_unitario,
            (oi.cantidad * oi.precio_unitario) AS subtotal
        FROM order_items oi
        JOIN products p ON p.id = oi.product_id
        WHERE oi.order_id = ?
        """,
        (order_id,)
    ).fetchall()

    return render_template("order_success.html", orden=orden, items=items, comprador=comprador)


# =====================================================
# EJECUCIÓN
# =====================================================
if __name__ == "__main__":
    os.makedirs(app.instance_path, exist_ok=True)
    app.run(debug=True)